"""A tela de seleção — que aqui é uma aba da própria planilha.

Resposta 13.5: das três opções, a escolhida foi a terceira — a planilha como
interface, com uma aba `Candidatos` de caixas de seleção. É a mais feia e a
mais provável de sobreviver a uma ausência prolongada de quem construiu: não
tem servidor para cair, dependência para atualizar, nem tela para aprender.
Quem usa já sabe usar Excel.

O fluxo é de três passos, e cada um é um comando só:

    1. abrir     lê as cotações e monta a aba `Candidatos`, com foto,
                 modelo, descrição, preço escolhido e MOQ
    2. conferir  lê as escolhas, calcula o custo e devolve a prévia na
                 própria aba, com os avisos — sem tocar no Funil
    3. gravar    escreve os marcados no `Funil` e na `Priorizacao`

O passo 2 existe por causa da seção 8: a prévia do custo tem que aparecer
**antes** de gravar. Ele não recalcula em fórmula do Excel de propósito —
duplicar o motor de custo em fórmula de planilha criaria duas contas para o
mesmo número, e um dia elas divergiriam.

Este módulo não decide nada de negócio: monta e lê a aba. Quem escolhe preço,
calcula m³ e custo é `normalizar/` e `custo/`, através de `mapeamento.py`.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from npd_tool.escrita.mapeamento import MARCAS, Celula, Complementos, preparar_linha
from npd_tool.escrita.ooxml import LinhaDeAba, Validacao, escrever_aba
from npd_tool.ingest.detector import ler_cotacao
from npd_tool.modelo import Ficha
from npd_tool.normalizar.embalagem import m3_por_unidade
from npd_tool.normalizar.nomes import nome_padronizado
from npd_tool.normalizar.precos import escolher_preco
from npd_tool.normalizar.specs import extrair_specs_eletricas

NOME_ABA = "Candidatos"

ULTIMA_LINHA_VALIDACAO = 400

ALTURA_LINHA = 60.0

# estilos que já existem no styles.xml da NPD (ver escrita/ooxml.py)
ESTILO_TITULO = "102"
ESTILO_TEXTO = "81"
ESTILO_CABECALHO = "97"
ESTILO_ROTULO = "103"

COL_MARCAR = "A"
COL_FOTO = "B"
COL_FORNECEDOR = "C"
COL_MODELO = "D"
COL_NOME = "E"
COL_DESCRICAO = "F"
COL_PRECO = "G"
COL_MOQ = "H"
COL_M3 = "I"
COL_NCM = "J"
COL_MARCA = "K"
COL_G1 = "L"
COL_CUSTO = "M"
COL_AVISOS = "N"
COL_ORIGEM = "O"

CABECALHOS = [
    (COL_MARCAR, "Marcar"),
    (COL_FOTO, "Foto"),
    (COL_FORNECEDOR, "Fornecedor"),
    (COL_MODELO, "Modelo"),
    (COL_NOME, "Nome do produto (editável)"),
    (COL_DESCRICAO, "Descrição"),
    (COL_PRECO, "Preço USD"),
    (COL_MOQ, "MOQ"),
    (COL_M3, "m³/un"),
    (COL_NCM, "NCM (preencher)"),
    (COL_MARCA, "Marca"),
    (COL_G1, "Sugestão G1"),
    (COL_CUSTO, "Custo estimado R$"),
    (COL_AVISOS, "Conferir"),
    (COL_ORIGEM, "origem — não editar"),
]

INSTRUCOES = [
    "COMO USAR — três passos, nesta ordem",
    "1. Marque com x na coluna A os produtos que quer levar para o Funil, e "
    "preencha o NCM (coluna J) e a Marca (coluna K) de cada um. Sem NCM o custo "
    "não é calculado e o produto entra com a coluna de custo vazia.",
    "2. Salve e feche esta planilha. No terminal, rode:   npd-tool conferir",
    "   O custo estimado aparece na coluna M desta aba, para você conferir antes.",
    "3. Se o custo fizer sentido, rode:   npd-tool gravar",
    "Esta aba é rascunho: é refeita toda vez que uma cotação nova é aberta, mas "
    "o que você marcou e digitou é preservado. Nada vai para o Funil antes do "
    "GRAVAR, e toda gravação faz backup antes.",
]

# derivadas das instruções, com uma linha em branco separando: mexer no texto
# acima não pode desalinhar o cabeçalho da tabela
LINHA_CABECALHO = len(INSTRUCOES) + 2
PRIMEIRA_LINHA = LINHA_CABECALHO + 1

LARGURAS = (
    '<cols><col min="1" max="1" width="8" customWidth="1"/>'
    '<col min="2" max="2" width="12" customWidth="1"/>'
    '<col min="3" max="3" width="22" customWidth="1"/>'
    '<col min="4" max="4" width="16" customWidth="1"/>'
    '<col min="5" max="5" width="42" customWidth="1"/>'
    '<col min="6" max="6" width="46" customWidth="1"/>'
    '<col min="7" max="9" width="11" customWidth="1"/>'
    '<col min="10" max="10" width="14" customWidth="1"/>'
    '<col min="11" max="11" width="14" customWidth="1"/>'
    '<col min="12" max="12" width="13" customWidth="1"/>'
    '<col min="13" max="13" width="17" customWidth="1"/>'
    '<col min="14" max="14" width="60" customWidth="1"/>'
    '<col min="15" max="15" width="30" customWidth="1"/></cols>'
)

MARCADORES = frozenset({"x", "X", "sim", "SIM", "s", "S", "1", "✓", "v", "V"})


@dataclass
class Candidato:
    """Uma linha da aba, com o suficiente para o gestor reconhecer o produto."""

    arquivo: Path
    indice: int
    ficha: Ficha
    nome_sugerido: str
    descricao_curta: str
    preco_usd: Decimal | None = None
    moq: int | None = None
    m3_unitario: Decimal | None = None
    sugestao_g1: str | None = None
    avisos: list[str] = field(default_factory=list)

    @property
    def identificador(self) -> str:
        return f"{self.arquivo}|{self.indice}"


@dataclass
class Selecao:
    """O que o gestor preencheu na aba, lido de volta."""

    linha: int
    arquivo: Path
    indice: int
    marcado: bool
    ncm: str | None
    marca: str | None
    nome: str | None


def _descricao_curta(ficha: Ficha, limite: int = 160) -> str:
    texto = " ".join((ficha.descricao_bruta or "").split())
    if len(texto) <= limite:
        return texto
    return texto[:limite].rsplit(" ", 1)[0] + "…"


def montar_candidatos(arquivos) -> list[Candidato]:
    """Lê as cotações e devolve os candidatos, sem tocar em planilha nenhuma."""
    candidatos: list[Candidato] = []
    for arquivo in arquivos:
        arquivo = Path(arquivo)
        for indice, ficha in enumerate(ler_cotacao(arquivo)):
            preco = escolher_preco(ficha)
            m3 = m3_por_unidade(ficha.embalagem)
            specs = extrair_specs_eletricas(ficha)
            nome, avisos_nome = nome_padronizado(ficha)

            candidatos.append(
                Candidato(
                    arquivo=arquivo,
                    indice=indice,
                    ficha=ficha,
                    nome_sugerido=nome,
                    descricao_curta=_descricao_curta(ficha),
                    preco_usd=preco.valor_final,
                    moq=preco.preco_base.moq if preco.preco_base else None,
                    m3_unitario=m3.valor,
                    sugestao_g1=specs.sugestao_g1,
                    avisos=[*avisos_nome, *preco.avisos, *m3.avisos, *specs.avisos],
                )
            )
    return candidatos


def _linha_de_candidato(
    candidato: Candidato,
    numero: int,
    custo: str | None = None,
    preenchido: "Selecao | None" = None,
) -> LinhaDeAba:
    """`preenchido` traz o que o gestor digitou numa passagem anterior.

    Reescrever a aba sem isso apagaria as marcações no passo `conferir` — a
    pessoa marcaria dez produtos, pediria a prévia e voltaria para uma aba em
    branco.
    """

    def texto(coluna, valor, estilo=ESTILO_TEXTO):
        return Celula(coluna, "texto", valor, estilo)

    celulas = [
        texto(COL_MARCAR, "x" if preenchido and preenchido.marcado else "", ESTILO_ROTULO),
        texto(COL_FORNECEDOR, candidato.ficha.fornecedor),
        texto(COL_MODELO, candidato.ficha.modelo or ""),
        texto(
            COL_NOME,
            (preenchido.nome if preenchido and preenchido.nome else candidato.nome_sugerido),
        ),
        texto(COL_DESCRICAO, candidato.descricao_curta),
        texto(COL_NCM, (preenchido.ncm if preenchido and preenchido.ncm else "")),
        texto(COL_MARCA, (preenchido.marca if preenchido and preenchido.marca else "")),
        texto(COL_G1, candidato.sugestao_g1 or ""),
        texto(COL_CUSTO, custo or ""),
        texto(COL_AVISOS, " | ".join(candidato.avisos)),
        texto(COL_ORIGEM, candidato.identificador),
    ]
    if candidato.preco_usd is not None:
        celulas.append(Celula(COL_PRECO, "numero", candidato.preco_usd, ESTILO_TEXTO))
    if candidato.moq is not None:
        celulas.append(Celula(COL_MOQ, "numero", candidato.moq, ESTILO_TEXTO))
    if candidato.m3_unitario is not None:
        celulas.append(Celula(COL_M3, "numero", candidato.m3_unitario, ESTILO_TEXTO))

    return LinhaDeAba(
        numero=numero,
        celulas=celulas,
        altura=ALTURA_LINHA,
        foto=candidato.ficha.foto,
        foto_formato=candidato.ficha.foto_formato or "png",
        coluna_foto=COL_FOTO,
    )


def montar_linhas(
    candidatos: list[Candidato],
    custos: dict[str, str] | None = None,
    preenchidos: dict[str, "Selecao"] | None = None,
) -> list[LinhaDeAba]:
    custos = custos or {}
    preenchidos = preenchidos or {}
    linhas = [
        LinhaDeAba(numero=1, celulas=[Celula("A", "texto", INSTRUCOES[0], ESTILO_TITULO)])
    ]
    for deslocamento, instrucao in enumerate(INSTRUCOES[1:], start=2):
        linhas.append(
            LinhaDeAba(numero=deslocamento, celulas=[Celula("A", "texto", instrucao, ESTILO_TEXTO)])
        )

    linhas.append(
        LinhaDeAba(
            numero=LINHA_CABECALHO,
            celulas=[
                Celula(coluna, "texto", titulo, ESTILO_CABECALHO)
                for coluna, titulo in CABECALHOS
            ],
        )
    )

    for deslocamento, candidato in enumerate(candidatos):
        numero = PRIMEIRA_LINHA + deslocamento
        chave = candidato.identificador
        linhas.append(
            _linha_de_candidato(
                candidato, numero, custos.get(chave), preenchidos.get(chave)
            )
        )
    return linhas


VALIDACOES = (
    Validacao(f"{COL_MARCAR}{PRIMEIRA_LINHA}:{COL_MARCAR}{ULTIMA_LINHA_VALIDACAO}", ("x",)),
    Validacao(f"{COL_MARCA}{PRIMEIRA_LINHA}:{COL_MARCA}{ULTIMA_LINHA_VALIDACAO}", MARCAS),
)


def escrever_aba_candidatos(
    npd_origem: Path,
    npd_destino: Path,
    candidatos: list[Candidato],
    custos: dict[str, str] | None = None,
    preenchidos: dict[str, "Selecao"] | None = None,
    com_backup: bool = True,
) -> dict:
    return escrever_aba(
        npd_origem,
        npd_destino,
        NOME_ABA,
        montar_linhas(candidatos, custos, preenchidos),
        colunas=LARGURAS,
        validacoes=VALIDACOES,
        congelar_ate=LINHA_CABECALHO,
        com_backup=com_backup,
    )


# ------------------------------------------------------------ leitura de volta

def ler_selecao(caminho_planilha: Path) -> list[Selecao]:
    """Lê o que o gestor marcou e preencheu na aba `Candidatos`."""
    import openpyxl

    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    try:
        if NOME_ABA not in wb.sheetnames:
            return []
        ws = wb[NOME_ABA]

        colunas = {coluna: indice + 1 for indice, (coluna, _) in enumerate(CABECALHOS)}
        selecoes = []
        for numero in range(PRIMEIRA_LINHA, ws.max_row + 1):
            origem = ws.cell(row=numero, column=colunas[COL_ORIGEM]).value
            if not origem or "|" not in str(origem):
                continue
            arquivo, _, indice = str(origem).rpartition("|")

            def valor(coluna):
                bruto = ws.cell(row=numero, column=colunas[coluna]).value
                texto = str(bruto).strip() if bruto is not None else ""
                return texto or None

            selecoes.append(
                Selecao(
                    linha=numero,
                    arquivo=Path(arquivo),
                    indice=int(indice),
                    marcado=(valor(COL_MARCAR) or "") in MARCADORES,
                    ncm=valor(COL_NCM),
                    marca=valor(COL_MARCA),
                    nome=valor(COL_NOME),
                )
            )
        return selecoes
    finally:
        wb.close()


class SelecaoInvalida(Exception):
    pass


def preparar_selecionados(selecoes, parametros, tabela_ncm, marca_padrao=None):
    """Das linhas marcadas para as linhas prontas do Funil.

    Relê a cotação de origem em vez de guardar a ficha: o que a aba carrega é o
    endereço do produto (arquivo + índice), não os dados dele. Assim o que for
    gravado é sempre o que está na cotação hoje, e não uma cópia que envelheceu
    entre um passo e outro.
    """
    marcadas = [s for s in selecoes if s.marcado]
    if not marcadas:
        return []

    cache: dict[Path, list[Ficha]] = {}
    preparadas = []
    for selecao in marcadas:
        if selecao.arquivo not in cache:
            if not selecao.arquivo.exists():
                raise SelecaoInvalida(
                    f"linha {selecao.linha}: a cotação {selecao.arquivo} não está "
                    "mais onde estava quando a aba foi montada"
                )
            cache[selecao.arquivo] = ler_cotacao(selecao.arquivo)
        fichas = cache[selecao.arquivo]
        if selecao.indice >= len(fichas):
            raise SelecaoInvalida(
                f"linha {selecao.linha}: a cotação {selecao.arquivo.name} mudou e "
                f"não tem mais o produto {selecao.indice}"
            )

        marca = selecao.marca or marca_padrao
        if marca not in MARCAS:
            raise SelecaoInvalida(
                f"linha {selecao.linha}: preencher a coluna Marca com "
                f"{' ou '.join(MARCAS)}"
            )

        preparadas.append(
            preparar_linha(
                fichas[selecao.indice],
                Complementos(marca=marca, ncm=selecao.ncm, nome=selecao.nome),
                parametros,
                tabela_ncm,
            )
        )
    return preparadas
