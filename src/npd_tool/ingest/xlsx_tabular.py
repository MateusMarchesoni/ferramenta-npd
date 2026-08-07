"""Parser de catálogos xlsx — caso real: Sunmile/Checky (`quotation(for Brazil).xlsx`).

O arquivo traz duas estruturas diferentes em duas abas, e o parser trata as
duas (PLANO.md seção 4.1, caso 'catálogo de 200 itens com 2 relevantes'):

**Aba em blocos** (`Sheet1`) — cabeçalho em algum lugar das primeiras linhas;
cada produto ocupa um bloco de várias linhas. A coluna do modelo acumula, de
cima para baixo, o modelo, a categoria e o preço; a descrição vem em linhas
sucessivas de outra coluna. Foto ancorada na coluna da esquerda.

**Aba em grupos de colunas** (`Sheet3`) — o mesmo catálogo repetido em quatro
grupos de três colunas, cada grupo com uma faixa de preço ligeiramente
diferente (mín, '-', máx). O que distingue os quatro grupos **não está escrito
em lugar nenhum do arquivo**, então todas as variantes são registradas e a
ficha sai com aviso, para o humano decidir. Ver PLANO.md seção 2.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import openpyxl

from npd_tool.ingest.comum import (
    extrair_dimensoes_mm,
    numeros_do_texto,
    para_decimal,
    texto_limpo,
)
from npd_tool.ingest.imagens import fotos_por_ancora_xlsx
from npd_tool.modelo import Embalagem, Ficha, Origem, Preco

PAPEIS_CABECALHO = {
    "model no": "modelo",
    "model no.": "modelo",
    "model": "modelo",
    "product picture": "foto",
    "picture": "foto",
    "description": "descricao",
    "certification": "certificado",
    "certificate": "certificado",
    "packing info": "embalagem",
    "packing": "embalagem",
    "delivery time": "entrega",
    "payment term": "pagamento",
}

# um modelo não tem espaços: 'DM-B', 'HM-350-220', 'BL-4X', 'CG-300'
_RE_MODELO = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-._/()+]*$")
_RE_PRECO_TEXTO = re.compile(r"^(?:usd|us\$|\$)\s*[\d.,]+$", re.IGNORECASE)
_MARCADORES_ACESSORIO = ("optional part", "optional", "accessory", "spare part")

# palavras que identificam a caixa de embarque — a única que serve para m³
_CHAVES_CARTON = ("ctn", "carton")
_CHAVES_EMBALAGEM_FRACA = ("packing",)
# caixa de presente e dimensão do produto nunca viram m³ (PLANO 6.4)
_CHAVES_IGNORAR = (
    "giftbox",
    "gift box",
    "box meas",
    "single box",
    "physical",
    "dimension",
    "product size",
    "unit size",
)


@dataclass
class _Bloco:
    linha_inicio: int
    linha_fim: int
    modelo: str
    categoria: str | None = None
    precos: list[tuple[Decimal, str]] = field(default_factory=list)
    descricao: list[str] = field(default_factory=list)
    certificacoes: list[str] = field(default_factory=list)
    embalagem_txt: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)


def _normalizar(texto) -> str:
    if texto is None:
        return ""
    return " ".join(str(texto).replace("\xa0", " ").split()).strip().lower().rstrip(":")


def _achar_cabecalho(ws, limite: int = 20) -> tuple[int, dict[str, int]] | None:
    for linha in range(1, min(limite, ws.max_row) + 1):
        mapa: dict[str, int] = {}
        for col in range(1, min(ws.max_column, 30) + 1):
            papel = PAPEIS_CABECALHO.get(_normalizar(ws.cell(row=linha, column=col).value))
            if papel and papel not in mapa:
                mapa[papel] = col
        if "modelo" in mapa:
            return linha, mapa
    return None


def _parece_modelo(valor) -> bool:
    if valor is None or isinstance(valor, (int, float, Decimal)):
        return False
    texto = str(valor).replace("\xa0", " ").strip()
    if not texto or len(texto) > 30:
        return False
    if _RE_PRECO_TEXTO.match(texto):
        return False
    return bool(_RE_MODELO.match(texto))


def _valor_de_preco(valor) -> Decimal | None:
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    if valor is None:
        return None
    texto = str(valor).replace("\xa0", " ").strip()
    if _RE_PRECO_TEXTO.match(texto):
        numeros = numeros_do_texto(texto)
        return numeros[0] if numeros else None
    return None


def _embalagem_de_textos(textos: list[str]) -> tuple[Embalagem, list[str]]:
    emb = Embalagem()
    avisos: list[str] = []
    for texto in textos:
        baixa = texto.lower()
        if any(chave in baixa for chave in _CHAVES_IGNORAR) and not any(
            chave in baixa for chave in _CHAVES_CARTON
        ):
            continue

        e_carton = any(chave in baixa for chave in _CHAVES_CARTON)
        e_packing = any(chave in baixa for chave in _CHAVES_EMBALAGEM_FRACA)
        if e_carton or e_packing:
            if emb.carton_mm is None:
                dims = extrair_dimensoes_mm(texto)
                if dims:
                    emb.carton_mm = dims
            m = re.search(r"(\d+)\s*pcs?\s*/\s*ctn", texto, re.IGNORECASE)
            if m and emb.pcs_por_carton is None:
                emb.pcs_por_carton = int(m.group(1))

        if ("n.w" in baixa or "n/w" in baixa) and emb.peso_liquido_kg is None:
            pesos = re.findall(r"(\d+(?:\.\d+)?)\s*kgs?", texto, re.IGNORECASE)
            if len(pesos) >= 1:
                emb.peso_liquido_kg = para_decimal(pesos[0])
            if len(pesos) >= 2:
                emb.peso_bruto_kg = para_decimal(pesos[1])

    if emb.carton_mm and emb.pcs_por_carton is None:
        avisos.append(
            "carton encontrado mas peças por caixa não informadas — m³ unitário não calculável"
        )
    return emb, avisos


def _blocos_da_aba(ws, linha_cabecalho: int, mapa: dict[str, int]) -> list[_Bloco]:
    col_modelo = mapa["modelo"]
    col_desc = mapa.get("descricao")
    col_cert = mapa.get("certificado")
    col_emb = mapa.get("embalagem")

    inicios: list[tuple[int, str]] = []
    anterior_preenchida = True
    for linha in range(linha_cabecalho + 1, ws.max_row + 1):
        valor = ws.cell(row=linha, column=col_modelo).value
        vazia = valor is None or not str(valor).strip()
        if not vazia and not anterior_preenchida and _parece_modelo(valor):
            inicios.append((linha, str(valor).strip()))
        anterior_preenchida = not vazia

    blocos: list[_Bloco] = []
    for i, (linha_inicio, modelo) in enumerate(inicios):
        fim = inicios[i + 1][0] - 1 if i + 1 < len(inicios) else ws.max_row
        bloco = _Bloco(linha_inicio=linha_inicio, linha_fim=fim, modelo=modelo)

        em_acessorio = False
        for linha in range(linha_inicio, fim + 1):
            bruto = ws.cell(row=linha, column=col_modelo).value
            texto = texto_limpo(bruto)

            if texto and any(m in texto.lower() for m in _MARCADORES_ACESSORIO):
                em_acessorio = True

            preco = _valor_de_preco(bruto)
            if preco is not None and linha != linha_inicio:
                rotulo = "acessório opcional" if em_acessorio else "padrão"
                bloco.precos.append((preco, rotulo))
            elif (
                texto
                and linha != linha_inicio
                and bloco.categoria is None
                and preco is None
                and not _parece_modelo(bruto)
            ):
                bloco.categoria = texto

            if col_desc:
                d = texto_limpo(ws.cell(row=linha, column=col_desc).value)
                if d:
                    bloco.descricao.append(d)
            if col_cert:
                c = texto_limpo(ws.cell(row=linha, column=col_cert).value)
                if c:
                    bloco.certificacoes.append(c)
            if col_emb:
                e = texto_limpo(ws.cell(row=linha, column=col_emb).value)
                if e:
                    bloco.embalagem_txt.append(e)

        if em_acessorio:
            bloco.avisos.append(
                "bloco contém acessório opcional cotado à parte — preço marcado como 'acessório opcional'"
            )
        blocos.append(bloco)
    return blocos


def _parse_aba_blocos(
    ws, nome_arquivo: str, fornecedor: str, contato: str | None
) -> list[Ficha]:
    achado = _achar_cabecalho(ws)
    if achado is None:
        return []
    linha_cabecalho, mapa = achado

    fotos = fotos_por_ancora_xlsx(ws)
    col_foto = mapa.get("foto", 1)
    fichas: list[Ficha] = []

    for bloco in _blocos_da_aba(ws, linha_cabecalho, mapa):
        origem = Origem(
            arquivo=nome_arquivo,
            aba_ou_pagina=ws.title,
            celula_ou_bbox=f"linhas {bloco.linha_inicio}-{bloco.linha_fim}",
            confianca="alta",
        )
        avisos = list(bloco.avisos)

        foto = foto_formato = None
        for (linha_ancora, coluna_ancora), (dados, formato) in sorted(fotos.items()):
            if coluna_ancora != col_foto:
                continue
            if bloco.linha_inicio <= linha_ancora <= bloco.linha_fim:
                foto, foto_formato = dados, formato
                break
        if foto is None:
            avisos.append("foto não encontrada")

        precos = [
            Preco(
                valor=valor,
                moeda="USD",
                incoterm=None,
                rotulo=rotulo,
                moq=None,
                origem=origem,
            )
            for valor, rotulo in bloco.precos
        ]
        if not precos:
            avisos.append("preço não informado na cotação")
        elif len([p for p in precos if p.rotulo == "padrão"]) > 1:
            avisos.append(
                "variantes de preço encontradas: "
                + ", ".join(f"{p.rotulo}={p.valor}" for p in precos)
            )

        embalagem, avisos_emb = _embalagem_de_textos(bloco.embalagem_txt)
        avisos.extend(avisos_emb)

        fichas.append(
            Ficha(
                fornecedor=fornecedor,
                contato=contato,
                data_cotacao=None,
                validade=None,
                modelo=bloco.modelo,
                descricao_bruta="\n".join(bloco.descricao),
                categoria=bloco.categoria,
                specs={},
                precos=precos,
                embalagem=embalagem,
                certificacoes=sorted(set(bloco.certificacoes)),
                foto=foto,
                foto_formato=foto_formato,
                origem=origem,
                avisos=avisos,
            )
        )
    return fichas


def _grupos_de_colunas(ws, linha_cabecalho: int) -> list[int]:
    colunas = []
    for col in range(1, ws.max_column + 1):
        if _normalizar(ws.cell(row=linha_cabecalho, column=col).value) in ("model no", "model no.", "model"):
            colunas.append(col)
    return colunas


def _parse_aba_grupos(
    ws, nome_arquivo: str, fornecedor: str, contato: str | None, linha_cabecalho: int
) -> list[Ficha]:
    grupos = _grupos_de_colunas(ws, linha_cabecalho)
    if len(grupos) < 2:
        return []

    principal = grupos[0]
    fichas: list[Ficha] = []
    linha = linha_cabecalho + 1

    while linha <= ws.max_row:
        valor = ws.cell(row=linha, column=principal).value
        if not _parece_modelo(valor):
            linha += 1
            continue

        modelo = str(valor).strip()
        categoria = texto_limpo(ws.cell(row=linha + 1, column=principal).value)

        origem = Origem(
            arquivo=nome_arquivo,
            aba_ou_pagina=ws.title,
            celula_ou_bbox=f"linha {linha}, grupos de colunas {grupos}",
            confianca="media",
        )

        precos: list[Preco] = []
        for indice_grupo, col in enumerate(grupos, start=1):
            for offset in range(1, 6):
                alvo = linha + offset
                if alvo > ws.max_row:
                    break
                valores = [
                    ws.cell(row=alvo, column=col + d).value for d in range(3)
                ]
                numeros = [
                    Decimal(str(v))
                    for v in valores
                    if isinstance(v, (int, float, Decimal))
                ]
                if numeros:
                    precos.append(
                        Preco(
                            valor=max(numeros),
                            moeda="USD",
                            incoterm=None,
                            rotulo=f"grupo {indice_grupo} (faixa {min(numeros)}–{max(numeros)})",
                            moq=None,
                            origem=origem,
                        )
                    )
                    break

        avisos = [
            "aba com quatro grupos de colunas de preço cujo significado não está "
            "documentado no arquivo — confirmar com o fornecedor antes de usar",
            "preço é faixa (mínimo–máximo); registrado o máximo de cada grupo",
        ]
        if not precos:
            avisos.append("preço não informado na cotação")

        fichas.append(
            Ficha(
                fornecedor=fornecedor,
                contato=contato,
                data_cotacao=None,
                validade=None,
                modelo=modelo,
                descricao_bruta=categoria or "",
                categoria=categoria,
                specs={},
                precos=precos,
                embalagem=Embalagem(),
                certificacoes=[],
                foto=None,
                foto_formato=None,
                origem=origem,
                avisos=avisos,
            )
        )
        linha += 1

    return fichas


def _identidade_fornecedor(ws) -> tuple[str, str | None]:
    fornecedor = "(fornecedor não identificado)"
    contato = None
    for linha in range(1, min(8, ws.max_row) + 1):
        for col in range(1, min(4, ws.max_column) + 1):
            texto = texto_limpo(ws.cell(row=linha, column=col).value)
            if not texto:
                continue
            if fornecedor.startswith("(") and re.search(
                r"\b(ltd|limited|ltda|industries|trading|co\.)\b", texto, re.IGNORECASE
            ):
                fornecedor = texto
            if contato is None and re.search(r"contact person", texto, re.IGNORECASE):
                contato = texto
    return fornecedor, contato


def parse_xlsx_tabular(caminho: Path) -> list[Ficha]:
    caminho = Path(caminho)
    wb = openpyxl.load_workbook(caminho, data_only=True)
    wb_imgs = openpyxl.load_workbook(caminho)

    fornecedor, contato = _identidade_fornecedor(wb[wb.sheetnames[0]])
    fichas: list[Ficha] = []

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        achado = _achar_cabecalho(ws)
        if achado is None:
            continue
        linha_cabecalho, mapa = achado

        if len(_grupos_de_colunas(ws, linha_cabecalho)) >= 2:
            fichas.extend(
                _parse_aba_grupos(ws, caminho.name, fornecedor, contato, linha_cabecalho)
            )
        else:
            ws_img = wb_imgs[nome_aba]
            ws_img_dados = ws
            # o parser lê valores de `ws` (data_only) e imagens de `ws_img`
            ws_img_dados._images = getattr(ws_img, "_images", [])
            fichas.extend(
                _parse_aba_blocos(ws_img_dados, caminho.name, fornecedor, contato)
            )

    return fichas
