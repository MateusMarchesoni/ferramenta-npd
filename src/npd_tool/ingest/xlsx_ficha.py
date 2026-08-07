"""Parser de ficha avulsa em xlsx — caso real: Milk Warmer da Frespro
(`Milk Warmer Estimate Quotation from Frespro--20260713.xlsx`).

Layout próprio e pequeno (PLANO.md seção 4.3): título, linha de datas,
uma linha de cabeçalho e uma linha por produto, seguidas de um bloco de
notas numeradas. Não há coluna de modelo — a identidade do produto é a
descrição, e o MOQ aparece solto no texto das notas.

Quando não há modelo, a ficha sai com a descrição no lugar dele e um aviso.
Inventar um código seria pior: o nome do produto é a chave que liga `Funil`
e `Priorizacao` (PLANO.md seção 3.5.1).
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from npd_tool.ingest.comum import (
    extrair_dimensoes_mm,
    numeros_do_texto,
    para_decimal,
    primeiro_inteiro,
    texto_limpo,
)
from npd_tool.ingest.imagens import fotos_por_ancora_xlsx
from npd_tool.modelo import Embalagem, Ficha, Origem, Preco

PAPEIS = {
    "description": "descricao",
    "product": "descricao",
    "model": "modelo",
    "model no": "modelo",
    "model no.": "modelo",
    "photo": "foto",
    "photo (just for reference)": "foto",
    "picture": "foto",
    "specification": "spec",
    "estimate quotation": "preco",
    "unit price": "preco",
    "price": "preco",
    "quotation": "preco",
    "moq": "moq",
}

_RE_NOTA = re.compile(r"^\s*(note|nota|\d+\s*\.)", re.IGNORECASE)


def _normalizar(valor) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).replace("\xa0", " ").split()).strip().lower().rstrip(":")


def _achar_cabecalho(ws, limite: int = 10) -> tuple[int, dict[int, str], dict[int, str]] | None:
    for linha in range(1, min(limite, ws.max_row) + 1):
        papel_por_col: dict[int, str] = {}
        rotulo_por_col: dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            bruto = ws.cell(row=linha, column=col).value
            chave = _normalizar(bruto)
            if not chave:
                continue
            rotulo_por_col[col] = str(bruto).strip()
            papel = PAPEIS.get(chave)
            if papel is None:
                papel = PAPEIS.get(re.sub(r"\s*\(.*?\)\s*", "", chave).strip())
            if papel:
                papel_por_col[col] = papel
        if "preco" in papel_por_col.values() and (
            "descricao" in papel_por_col.values() or "modelo" in papel_por_col.values()
        ):
            return linha, papel_por_col, rotulo_por_col
    return None


def _e_linha_de_nota(ws, linha: int, max_col: int) -> bool:
    primeiro = texto_limpo(ws.cell(row=linha, column=1).value)
    if not primeiro:
        return False
    resto_vazio = all(
        ws.cell(row=linha, column=c).value is None for c in range(2, max_col + 1)
    )
    return resto_vazio and bool(_RE_NOTA.match(primeiro))


def _datas(ws) -> tuple[date | None, date | None]:
    for linha in range(1, min(4, ws.max_row) + 1):
        texto = texto_limpo(ws.cell(row=linha, column=1).value)
        if not texto or "date" not in texto.lower():
            continue
        datas = re.findall(r"\d{4}-\d{2}-\d{2}", texto)
        conv = []
        for d in datas:
            try:
                conv.append(datetime.strptime(d, "%Y-%m-%d").date())
            except ValueError:
                pass
        return (
            conv[0] if len(conv) >= 1 else None,
            conv[1] if len(conv) >= 2 else None,
        )
    return None, None


def parse_xlsx_ficha(caminho: Path) -> list[Ficha]:
    caminho = Path(caminho)
    wb = openpyxl.load_workbook(caminho, data_only=True)
    wb_img = openpyxl.load_workbook(caminho)
    fichas: list[Ficha] = []

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        achado = _achar_cabecalho(ws)
        if achado is None:
            continue
        linha_cabecalho, papel_por_col, rotulo_por_col = achado

        data_cotacao, validade = _datas(ws)
        fotos = fotos_por_ancora_xlsx(wb_img[nome_aba])

        notas: list[str] = []
        for linha in range(linha_cabecalho + 1, ws.max_row + 1):
            if _e_linha_de_nota(ws, linha, ws.max_column):
                texto = texto_limpo(ws.cell(row=linha, column=1).value)
                if texto:
                    notas.append(texto)
        moq_das_notas = None
        for nota in notas:
            m = re.search(r"moq\D{0,40}(\d[\d,]*)", nota, re.IGNORECASE)
            if m:
                moq_das_notas = primeiro_inteiro(m.group(1))
                break

        for linha in range(linha_cabecalho + 1, ws.max_row + 1):
            if _e_linha_de_nota(ws, linha, ws.max_column):
                continue
            valores = {
                col: texto_limpo(ws.cell(row=linha, column=col).value)
                for col in rotulo_por_col
            }
            if not any(valores.values()):
                continue

            def por_papel(papel: str) -> str | None:
                for col, p in papel_por_col.items():
                    if p == papel and valores.get(col):
                        return valores[col]
                return None

            descricao = por_papel("descricao")
            modelo = por_papel("modelo")
            preco_texto = por_papel("preco")
            if not (descricao or modelo) or not preco_texto:
                continue

            avisos: list[str] = []
            if not modelo:
                modelo = descricao
                avisos.append(
                    "cotação não traz código de modelo — usada a descrição como identificação"
                )

            specs = {}
            for col, rotulo in rotulo_por_col.items():
                if papel_por_col.get(col) in ("descricao", "modelo", "preco", "foto"):
                    continue
                if valores.get(col):
                    specs[rotulo] = valores[col]

            origem = Origem(
                arquivo=caminho.name,
                aba_ou_pagina=nome_aba,
                celula_ou_bbox=f"linha {linha}",
                confianca="alta",
            )

            numeros = numeros_do_texto(preco_texto)
            precos = []
            if numeros:
                precos.append(
                    Preco(
                        valor=numeros[0],
                        moeda="USD",
                        incoterm=None,
                        rotulo="padrão",
                        moq=primeiro_inteiro(por_papel("moq")) or moq_das_notas,
                        origem=origem,
                    )
                )
            else:
                avisos.append("preço não informado na cotação")

            foto = foto_formato = None
            for (linha_ancora, coluna_ancora), (dados, formato) in fotos.items():
                if linha_ancora == linha and papel_por_col.get(coluna_ancora) == "foto":
                    foto, foto_formato = dados, formato
                    break
            if foto is None:
                avisos.append("foto não encontrada")

            embalagem = Embalagem()
            for rotulo, valor in specs.items():
                baixa = rotulo.lower()
                if any(c in baixa for c in ("carton", "ctn", "packing")):
                    embalagem.carton_mm = embalagem.carton_mm or extrair_dimensoes_mm(valor)
                if "net weight" in baixa or "n.w" in baixa:
                    embalagem.peso_liquido_kg = embalagem.peso_liquido_kg or para_decimal(
                        (numeros_do_texto(valor) or [None])[0]
                    )
            if embalagem.carton_mm is None:
                avisos.append(
                    "cotação não traz dimensões de caixa de embarque — m³ unitário não calculável"
                )

            if notas:
                avisos.append("notas da cotação: " + " | ".join(notas))

            fichas.append(
                Ficha(
                    fornecedor="Frespro",
                    contato=None,
                    data_cotacao=data_cotacao,
                    validade=validade,
                    modelo=modelo,
                    descricao_bruta=descricao or "",
                    categoria=None,
                    specs=specs,
                    precos=precos,
                    embalagem=embalagem,
                    certificacoes=[],
                    foto=foto,
                    foto_formato=foto_formato,
                    origem=origem,
                    avisos=avisos,
                )
            )

    return fichas
