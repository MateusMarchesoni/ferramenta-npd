"""Identifica o formato de uma cotação e roteia para o parser certo.

A decisão é estrutural, nunca pelo nome do arquivo: fornecedores renomeiam,
reencaminham e reexportam, e um nome de arquivo não é contrato.

O que separa os dois layouts de xlsx é a **orientação dos rótulos**:

- na ficha transposta (Frespro), a coluna da esquerda é uma pilha de rótulos
  de atributo — `Image`, `Category`, `Model No.`, `General Specification`,
  `Quotation` — e cada coluna à direita é um produto;
- no catálogo tabular (Sunmile), os mesmos rótulos aparecem lado a lado numa
  única linha de cabeçalho, e cada linha abaixo é um produto.

Contar rótulos conhecidos por linha e por coluna distingue os dois sem
depender de nenhum rótulo específico estar presente — que é o que quebrava
quando `Model No` aparecia nos dois formatos.

**Nenhum arquivo é recusado por não casar com esta lista.** O vocabulário
(`rotulos.py`) decide qual parser é o *melhor*, não se existe algum: quando
nada é reconhecido, a cotação vai para `xlsx_generico.py`, que lê pela forma da
planilha e devolve fichas de confiança baixa, com aviso. Recusar o arquivo
devolve à pessoa um problema que ela não tem como resolver; devolver o que deu
para ler, marcado como incerto, ela confere na tela.
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable

import openpyxl

from npd_tool.ingest.rotulos import PAPEIS_DE_IDENTIDADE, e_rotulo, papel_de
from npd_tool.modelo import Ficha

FormatoDesconhecido = "desconhecido"

# um cabeçalho de verdade tem mais de um rótulo; um único acerto é ruído
MINIMO_PARA_CABECALHO = 2

# A NPD tem colunas chamadas `Produto`, `Marca` e `Categoria`, que são rótulos
# de cotação legítimos — com o vocabulário largo, ela se parece o suficiente
# com um catálogo para ser lida como um, e despejaria trezentos "produtos" de
# volta na lista de candidatos. Ela se identifica pelas próprias abas.
ABAS_DA_NPD = frozenset({"Funil", "Pesos", "Priorizacao"})


class FormatoNaoSuportado(Exception):
    pass


def _e_rotulo(valor) -> bool:
    return e_rotulo(valor)


def _densidades(ws, ate_linha: int = 15, ate_coluna: int = 30) -> tuple[int, int]:
    """(maior nº de rótulos numa linha, maior nº de rótulos numa coluna)."""
    max_linha = min(ate_linha, ws.max_row)
    max_col = min(ate_coluna, ws.max_column)

    por_linha = [0] * (max_linha + 1)
    por_coluna = [0] * (max_col + 1)

    for linha in range(1, max_linha + 1):
        for col in range(1, max_col + 1):
            if _e_rotulo(ws.cell(row=linha, column=col).value):
                por_linha[linha] += 1
                por_coluna[col] += 1

    return max(por_linha, default=0), max(por_coluna, default=0)


def _tem_coluna_de_modelo(ws, ate_linha: int = 15, ate_coluna: int = 30) -> bool:
    """Uma coluna que identifica o produto — `Model No.`, `Item No.`, `Produto`.

    É o que separa um catálogo (uma linha por produto) de uma ficha avulsa (um
    produto só, com os atributos espalhados).
    """
    for linha in range(1, min(ate_linha, ws.max_row) + 1):
        for col in range(1, min(ate_coluna, ws.max_column) + 1):
            if papel_de(ws.cell(row=linha, column=col).value) in PAPEIS_DE_IDENTIDADE:
                return True
    return False


def detectar_formato(caminho: Path) -> str:
    caminho = Path(caminho)
    sufixo = caminho.suffix.lower()

    if sufixo == ".pdf":
        return "pdf_tabular"

    if sufixo == ".xls":
        # o .xls é um formato binário de 1997 que o openpyxl não abre. Dizer
        # "não suportado" deixa a pessoa sem saída; dizer como converter
        # resolve em dois cliques, e o Excel dela já faz isso.
        raise FormatoNaoSuportado(
            f"{caminho.name} está no formato antigo .xls, que a ferramenta não "
            "abre. Abra o arquivo no Excel e use Arquivo → Salvar como → "
            "Pasta de Trabalho do Excel (.xlsx); depois adicione o .xlsx aqui."
        )

    if sufixo not in (".xlsx", ".xlsm"):
        raise FormatoNaoSuportado(
            f"{caminho.name}: a ferramenta lê cotação em .xlsx e em .pdf "
            f"(este arquivo é {sufixo or 'sem extensão'})"
        )

    wb = openpyxl.load_workbook(caminho, data_only=True)
    try:
        if ABAS_DA_NPD <= set(wb.sheetnames):
            raise FormatoNaoSuportado(
                f"{caminho.name} é a própria planilha NPD (tem as abas Funil, "
                "Pesos e Priorizacao), não uma cotação de fornecedor. Escolha "
                "os arquivos que os fornecedores enviaram."
            )

        melhor_linha = melhor_coluna = 0
        tem_modelo = False
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            por_linha, por_coluna = _densidades(ws)
            melhor_linha = max(melhor_linha, por_linha)
            melhor_coluna = max(melhor_coluna, por_coluna)
            tem_modelo = tem_modelo or _tem_coluna_de_modelo(ws)

        if max(melhor_linha, melhor_coluna) < MINIMO_PARA_CABECALHO:
            return FormatoDesconhecido

        if melhor_coluna > melhor_linha:
            return "xlsx_transposto"

        # tabular: com coluna de modelo é catálogo; sem ela, ficha avulsa
        return "xlsx_tabular" if tem_modelo else "xlsx_ficha"
    finally:
        wb.close()


def _parsers() -> dict[str, Callable[[Path], list[Ficha]]]:
    from npd_tool.ingest.pdf_tabular import parse_pdf_tabular
    from npd_tool.ingest.xlsx_ficha import parse_xlsx_ficha
    from npd_tool.ingest.xlsx_tabular import parse_xlsx_tabular
    from npd_tool.ingest.xlsx_transposto import parse_xlsx_transposto

    return {
        "pdf_tabular": parse_pdf_tabular,
        "xlsx_transposto": parse_xlsx_transposto,
        "xlsx_tabular": parse_xlsx_tabular,
        "xlsx_ficha": parse_xlsx_ficha,
    }


def ler_cotacao(caminho: Path) -> list[Ficha]:
    """Lê uma cotação e devolve as fichas — sem recusar layout desconhecido.

    A cotação de um fornecedor novo raramente é igual às que existiam quando o
    parser foi escrito, e a pessoa que a recebeu não tem como reformatá-la. Por
    isso a leitura tem dois níveis: o parser específico, quando o layout é
    reconhecido; e o genérico, que lê pela forma da planilha e marca tudo como
    confiança baixa. Só sobra erro quando nem a forma se sustenta — e aí o erro
    diz o que fazer.
    """
    caminho = Path(caminho)
    formato = detectar_formato(caminho)
    parser = _parsers().get(formato)

    fichas: list[Ficha] = []
    if parser is not None:
        fichas = parser(caminho)
        if fichas:
            return fichas

    if caminho.suffix.lower() == ".pdf":
        # o PDF já passou pelo pdf_tabular, que é o único leitor de PDF que
        # existe: não há segundo nível para tentar
        raise FormatoNaoSuportado(
            f"{caminho.name}: abri o PDF mas não encontrei nenhuma tabela de "
            "produtos nele. PDF de catálogo em imagem (página escaneada) não "
            "tem texto para ler — nesse caso peça a versão em Excel ao "
            "fornecedor, ou lance o produto à mão."
        )

    from npd_tool.ingest.xlsx_generico import parse_xlsx_generico

    fichas = parse_xlsx_generico(caminho)
    if fichas:
        return fichas

    raise FormatoNaoSuportado(
        f"{caminho.name}: abri a planilha mas não encontrei produtos nela. "
        "A ferramenta procura uma linha ou coluna de cabeçalho (com algo como "
        "Model/Item/Produto, Price/Preço, MOQ, Packing) e, se não achar, "
        "procura uma tabela com códigos e preços. Confira se a cotação está na "
        "primeira aba e se as colunas têm título."
    )
