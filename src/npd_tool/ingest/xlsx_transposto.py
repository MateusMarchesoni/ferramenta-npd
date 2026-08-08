"""Parser do formato 'ficha transposta' — uma aba por família de produto,
cada produto numa coluna, cada atributo numa linha. Caso real: Frespro.

Ver PLANO.md seção 4.2. Estrutura recorrente por aba:
  linha 1  título
  linha 2  data da cotação e validade
  linha 3  "Image" — fotos ancoradas, uma por coluna de produto
  linha 4  "Category"
  linha 5  "Model No." — define quais colunas são produtos
  linha 6+ blocos de specs, cada um terminando no bloco "Quotation"
  bloco Quotation: uma ou mais linhas "Unit Price..." e uma linha "MOQ"

**As linhas são procuradas pelo rótulo, não pelo número.** A estrutura acima é
a da Frespro, e continua valendo como último recurso; mas outra fábrica manda a
mesma ficha com o cabeçalho duas linhas acima, e ler a linha 5 às cegas fazia a
especificação técnica virar o nome do modelo — um produto no funil chamado
"2.2L bowl, 550W, 2 speeds". Procurar `Model No.` na coluna da esquerda é o
mesmo trabalho que o parser tabular faz nas colunas, com o mesmo vocabulário
(`rotulos.py`).

Regra do maior valor (seção 4.4): quando há mais de um preço para o mesmo
produto (com/sem acessório, faixa de MOQ, montado/SKD), a ficha guarda
todos em `precos`, mas normalizar/precos.py escolhe o maior. Aqui só
registramos as variantes.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from npd_tool.ingest.comum import (
    extrair_dimensoes_mm,
    primeiro_inteiro,
    texto_limpo,
)
from npd_tool.ingest.rotulos import papel_de
from npd_tool.modelo import Embalagem, Ficha, Origem, Preco

# As posições da ficha da Frespro, que foi a cotação a partir da qual este
# parser foi escrito. Elas continuam valendo como **último recurso**: quando o
# rótulo da linha não é reconhecido, a ficha do fornecedor conhecido ainda é
# lida. O que não vale mais é *começar* por elas — outra fábrica usa a mesma
# estrutura com o cabeçalho duas linhas acima, e ler a linha 5 às cegas fazia a
# especificação técnica virar o nome do modelo.
LINHA_IMAGEM = 3
LINHA_CATEGORIA = 4
LINHA_MODELO = 5
LIMITE_DE_BUSCA = 12
ROTULOS_SECAO_VAZIA = {"general specification", "features"}


def _linha_do_papel(ws, papel: str, padrao: int | None) -> int | None:
    """A linha cujo rótulo (coluna A) tem este papel, ou `padrao`.

    A ficha transposta identifica cada linha pelo rótulo à esquerda; procurar
    o rótulo é o mesmo trabalho que o parser tabular faz nas colunas, e é o que
    torna o parser independente de qual fornecedor mandou o arquivo.
    """
    for linha in range(1, min(LIMITE_DE_BUSCA, ws.max_row) + 1):
        if papel_de(ws.cell(row=linha, column=1).value) == papel:
            # uma linha de rótulo só vale se houver produto à direita dela
            if any(
                texto_limpo(ws.cell(row=linha, column=col).value)
                for col in range(2, min(ws.max_column, 12) + 1)
            ):
                return linha
    return padrao


@dataclass
class _Produto:
    coluna: int
    modelo: str
    specs: dict[str, str]
    precos_brutos: list[tuple[str, Decimal, str | None]]  # (rotulo, valor, moq_texto)
    foto: bytes | None
    foto_formato: str | None


def _categoria_do_titulo(titulo) -> str | None:
    """'Hot Shot Steamer  Quotation' -> 'Hot Shot Steamer'."""
    if titulo is None:
        return None
    texto = " ".join(str(titulo).replace("\xa0", " ").split())
    texto = re.sub(r"\s*quotation\s*$", "", texto, flags=re.IGNORECASE).strip()
    return texto or None


def _parse_data_validade(texto: str | None) -> tuple[date | None, date | None]:
    if not texto:
        return None, None
    datas = re.findall(r"\d{4}-\d{2}-\d{2}", texto)
    cotacao = _parse_date(datas[0]) if len(datas) >= 1 else None
    validade = _parse_date(datas[1]) if len(datas) >= 2 else None
    return cotacao, validade


def _parse_date(texto: str) -> date | None:
    try:
        return datetime.strptime(texto, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_decimal(valor) -> Decimal | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip()
    texto = re.sub(r"[^\d.,-]", "", texto)
    texto = texto.replace(",", "")
    if not texto:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _parse_moq(valor) -> int | None:
    if valor is None:
        return None
    m = re.search(r"\d+", str(valor))
    return int(m.group(0)) if m else None


def _embalagem_da_coluna(itens: list) -> Embalagem:
    """A embalagem de um produto, a partir das linhas de rótulo já classificadas.

    Só o que o fornecedor escreveu na linha certa entra aqui — `Carton Size`
    vira medida de embarque, `Product Size` nem chega a este ponto, porque o
    vocabulário já lhe deu outro papel. É a mesma regra do parser tabular, pelo
    mesmo motivo: a caixa errada produz um m³ errado que ninguém confere.
    """
    emb = Embalagem()
    for papel, valor in itens:
        texto = str(valor).strip()
        if papel == "embalagem" and emb.carton_mm is None:
            emb.carton_mm = extrair_dimensoes_mm(texto)
        elif papel == "pcs_por_caixa" and emb.pcs_por_carton is None:
            emb.pcs_por_carton = primeiro_inteiro(texto)
        elif papel == "cbm" and emb.cbm_total is None:
            cbm = _parse_decimal(valor)
            if cbm is not None and cbm > 0:
                emb.cbm_total = cbm
        elif papel == "peso_liquido" and emb.peso_liquido_kg is None:
            emb.peso_liquido_kg = _parse_decimal(valor)
        elif papel == "peso_bruto" and emb.peso_bruto_kg is None:
            emb.peso_bruto_kg = _parse_decimal(valor)
    if emb.cbm_total is not None and emb.qty_referencia is None:
        emb.qty_referencia = emb.pcs_por_carton or 1
    return emb


def _colunas_produto(ws, linha_modelo: int) -> dict[int, str]:
    colunas: dict[int, str] = {}
    for cell in ws[linha_modelo]:
        if cell.column == 1:
            continue
        if cell.value is not None and str(cell.value).strip():
            colunas[cell.column] = str(cell.value).strip()
    return colunas


def _fotos_por_coluna(ws, linha_imagem: int) -> dict[int, tuple[bytes, str]]:
    fotos: dict[int, tuple[bytes, str]] = {}
    for img in getattr(ws, "_images", []):
        frm = img.anchor._from
        linha_1based = frm.row + 1
        coluna_1based = frm.col + 1
        if linha_1based != linha_imagem or coluna_1based == 1:
            continue
        fotos[coluna_1based] = (img._data(), img.format)
    return fotos


def _linha_e_so_rotulo(row) -> bool:
    """Linha com texto só na coluna A — cabeçalho de seção como
    'General Specification', não um atributo de verdade."""
    rotulo = row[0].value
    if rotulo is None:
        return False
    resto = [c.value for c in row[1:]]
    return str(rotulo).strip().lower() in ROTULOS_SECAO_VAZIA and all(
        v is None for v in resto
    )


def parse_xlsx_transposto(caminho: Path) -> list[Ficha]:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    fichas: list[Ficha] = []

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        linha_modelo = _linha_do_papel(ws, "modelo", None)
        if linha_modelo is None:
            linha_modelo = _linha_do_papel(ws, "nome", LINHA_MODELO)
        linha_categoria = _linha_do_papel(ws, "categoria", LINHA_CATEGORIA)
        linha_imagem = _linha_do_papel(ws, "foto", LINHA_IMAGEM)
        if ws.max_row < linha_modelo:
            continue

        colunas_produto = _colunas_produto(ws, linha_modelo)
        if not colunas_produto:
            continue

        data_cotacao, validade = _parse_data_validade(
            ws.cell(row=2, column=1).value
        )
        categoria_por_coluna = {
            col: str(ws.cell(row=linha_categoria, column=col).value).strip()
            for col in colunas_produto
            if ws.cell(row=linha_categoria, column=col).value is not None
        }
        # algumas abas deixam `Category` em branco; o título da aba
        # ('Hot Shot Steamer Quotation') diz a mesma coisa e está no arquivo
        titulo_aba = _categoria_do_titulo(ws.cell(row=1, column=1).value)
        fotos = _fotos_por_coluna(ws, linha_imagem)

        produtos = {
            col: _Produto(
                coluna=col,
                modelo=modelo,
                specs={},
                precos_brutos=[],
                foto=fotos.get(col, (None, None))[0],
                foto_formato=fotos.get(col, (None, None))[1],
            )
            for col, modelo in colunas_produto.items()
        }

        ultimo_rotulo: str | None = None
        em_quotation = False
        moq_por_coluna: dict[int, str] = {}
        embalagem_por_coluna: dict[int, list] = {col: [] for col in produtos}

        for row in ws.iter_rows(min_row=linha_modelo + 1, max_row=ws.max_row):
            rotulo_celula = row[0].value
            rotulo = str(rotulo_celula).strip() if rotulo_celula is not None else None
            papel = papel_de(rotulo) if rotulo else None

            def _valores():
                for col in produtos:
                    valor = row[col - 1].value if col - 1 < len(row) else None
                    if valor is not None and str(valor).strip():
                        yield col, valor

            if papel == "preco":
                valores = list(_valores())
                if valores:
                    # o preço na própria linha do rótulo: `Quotation | 74 | 118`
                    variante = re.sub(
                        r"^\s*(unit\s+price|quotation|pre[çc]o)", "", rotulo,
                        flags=re.IGNORECASE,
                    ).strip(" -:") or "padrão"
                    for col, valor in valores:
                        decimal = _parse_decimal(valor)
                        if decimal is not None:
                            produtos[col].precos_brutos.append((variante, decimal, None))
                    ultimo_rotulo = None
                    continue
                # rótulo de preço sem número é o **título** do bloco de preços,
                # e as linhas seguintes é que trazem as variantes
                em_quotation = True
                ultimo_rotulo = None
                continue

            if papel == "moq":
                for col, valor in _valores():
                    moq_por_coluna[col] = str(valor)
                ultimo_rotulo = None
                continue

            if papel in ("embalagem", "cbm", "pcs_por_caixa", "peso_liquido",
                         "peso_bruto"):
                for col, valor in _valores():
                    embalagem_por_coluna[col].append((papel, valor))
                    # continua valendo como spec: a linha de embalagem é parte
                    # da ficha técnica que a pessoa lê na tela, e tirá-la de lá
                    # esconderia informação que o fornecedor escreveu
                    produtos[col].specs[rotulo] = str(valor).strip()
                ultimo_rotulo = None
                continue

            if rotulo and _linha_e_so_rotulo(row):
                ultimo_rotulo = None
                continue

            # depois do bloco de preços vêm as condições gerais e o texto de
            # marketing, que não são atributo de produto nenhum
            if em_quotation:
                break

            if rotulo is not None:
                ultimo_rotulo = rotulo
                for col, produto in produtos.items():
                    valor = row[col - 1].value if col - 1 < len(row) else None
                    if valor is not None:
                        produto.specs[rotulo] = str(valor).strip()
            elif ultimo_rotulo is not None:
                for col, produto in produtos.items():
                    valor = row[col - 1].value if col - 1 < len(row) else None
                    if valor is not None and ultimo_rotulo in produto.specs:
                        produto.specs[ultimo_rotulo] += "; " + str(valor).strip()
                    elif valor is not None:
                        produto.specs[ultimo_rotulo] = str(valor).strip()

        for col, produto in produtos.items():
            avisos: list[str] = []
            categoria = categoria_por_coluna.get(col)
            if not categoria and titulo_aba:
                categoria = titulo_aba
                avisos.append(
                    f"aba não preenche 'Category' — categoria '{titulo_aba}' "
                    "deduzida do título da aba"
                )
            origem = Origem(
                arquivo=Path(caminho).name,
                aba_ou_pagina=nome_aba,
                celula_ou_bbox=f"col {col}",
                confianca="alta",
            )
            precos = [
                Preco(
                    valor=valor,
                    moeda="USD",
                    incoterm=None,
                    rotulo=rotulo,
                    moq=_parse_moq(moq_por_coluna.get(col)),
                    origem=origem,
                )
                for rotulo, valor, _ in produto.precos_brutos
            ]
            if len(precos) > 1:
                rotulos = ", ".join(f"{p.rotulo}={p.valor}" for p in precos)
                avisos.append(f"variantes de preço encontradas: {rotulos}")
            if not precos:
                avisos.append("nenhum preço encontrado")
            if produto.foto is None:
                avisos.append("foto não encontrada")

            embalagem = _embalagem_da_coluna(embalagem_por_coluna.get(col, []))
            if (
                embalagem.carton_mm
                and embalagem.pcs_por_carton is None
                and embalagem.cbm_total is None
            ):
                # sem peças por caixa não há m³ unitário, e supor uma peça por
                # caixa seria inventar o número que multiplica o frete
                avisos.append(
                    "medida de caixa encontrada mas peças por caixa não — m³ "
                    "unitário não calculável"
                )

            fichas.append(
                Ficha(
                    fornecedor="Frespro",
                    contato=None,
                    data_cotacao=data_cotacao,
                    validade=validade,
                    modelo=produto.modelo,
                    descricao_bruta=categoria or "",
                    categoria=categoria,
                    specs=produto.specs,
                    precos=precos,
                    embalagem=embalagem,
                    certificacoes=[],
                    foto=produto.foto,
                    foto_formato=produto.foto_formato,
                    origem=origem,
                    avisos=avisos,
                )
            )

    return fichas
