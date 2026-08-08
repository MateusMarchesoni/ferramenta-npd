"""Último nível de leitura: a planilha cujo layout ninguém previu.

Os outros parsers conhecem um layout cada, aprendido de uma cotação real. Este
não conhece nenhum. Ele parte de uma observação que vale para praticamente toda
cotação de fornecedor, em qualquer idioma: **existe uma região retangular onde
cada linha é um produto**, e dentro dela existe uma coluna que identifica o
produto e uma coluna de números que são preços.

Achar essa região é um problema de forma, não de vocabulário:

1. a **região de dados** é o maior trecho de linhas seguidas com pelo menos
   duas células preenchidas — títulos, logotipos e blocos de observação ficam
   de fora porque são linhas esparsas;
2. a **coluna de preço** é a coluna com mais números numa faixa plausível,
   preferindo a que tem casas decimais — quantidade e ano são inteiros redondos,
   preço quase nunca é;
3. a **coluna de identidade** é a coluna de texto curto com mais valores
   distintos — código de modelo varia a cada linha, categoria se repete.

Nada disso é confiável do jeito que `Model No.` é confiável, e o módulo não
finge que é: toda ficha daqui sai com `confianca="baixa"` e um aviso dizendo
quais colunas foram adivinhadas, para a pessoa conferir na tela antes de gravar.
O contrato de nunca inventar dado continua valendo — o que é chute é o
*significado da coluna*, nunca o valor da célula, que é copiado como está.

A alternativa a este módulo não é uma leitura melhor: é a mensagem "nenhum
produto foi reconhecido", que devolve à pessoa um problema que ela não pode
resolver.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from npd_tool.ingest.cabecalho import (
    fim_do_cabecalho,
    preencher_mesclas_verticais,
    rotulos_achatados,
)
from npd_tool.ingest.comum import extrair_dimensoes_mm, primeiro_inteiro, texto_limpo
from npd_tool.ingest.imagens import fotos_por_ancora_xlsx
from npd_tool.ingest.rotulos import faixa_de_quantidade, mapear_papeis, mapear_todos
from npd_tool.modelo import Embalagem, Ficha, Origem, Preco

# Limites de varredura. Uma cotação com mais de 40 colunas ou 5.000 linhas não
# é uma cotação, é um relatório de sistema — e varrer tudo custa minutos.
MAX_COLUNAS = 40
MAX_LINHAS = 5000

# Faixa em que um número pode ser preço de equipamento de cozinha em USD. Fora
# dela quase sempre é peso, dimensão, ano ou código.
PRECO_MINIMO = Decimal("0.5")
PRECO_MAXIMO = Decimal("500000")

# uma coluna precisa de massa para ser eleita; duas células não são padrão
MINIMO_DE_LINHAS_NA_REGIAO = 2
MINIMO_PARA_ELEGER_COLUNA = 2

# sem nenhum rótulo reconhecido, metade da coluna de identidade precisa parecer
# código de modelo para a planilha ser aceita como cotação
FRACAO_MINIMA_DE_CODIGO = 0.5

# acima disto a leitura quase certamente pegou a planilha errada
MAXIMO_DE_FICHAS = 500

# um código de modelo não tem espaço: 'HY-201', 'BT.500', 'CG300/A'. É o que
# separa a coluna de código da coluna de nome do produto, que também é texto
# curto e também muda a cada linha.
_RE_CODIGO = re.compile(r"^(?=.*\d)[A-Za-z0-9][A-Za-z0-9\-._/+]{1,29}$")
_RE_SO_NUMERO = re.compile(r"^\d+([.,]\d+)?$")
_RE_MOEDA = re.compile(r"(usd|us\$|r\$|rmb|cny|eur|\$|€|¥)", re.IGNORECASE)
_RE_BR = re.compile(r"^\d{1,3}(\.\d{3})+,\d+$")
_RE_US = re.compile(r"^\d{1,3}(,\d{3})+(\.\d+)?$")
_RE_VIRGULA_DECIMAL = re.compile(r"^\d+,\d{1,2}$")
# a unidade colada no número: 'USD39.00/set', '45.00 /pc', '12,50/un'
_RE_UNIDADE_DE_PRECO = re.compile(
    r"/\s*(pcs?|pieces?|sets?|units?|un|unid|unidades?|ea|each|ctn|carton|"
    r"cx|caixa|kg|pe[çc]as?)\.?$",
    re.IGNORECASE,
)

# linhas que são rodapé de cotação, não produto
_RE_RODAPE = re.compile(
    r"^\s*(note|notes|nota|notas|remark|remarks|obs|total|subtotal|payment|"
    r"delivery|validity|valid|bank|contact|tel|e-?mail|address)\b",
    re.IGNORECASE,
)


def para_preco(valor) -> Decimal | None:
    """Número de célula ou de texto, nas duas convenções decimais.

    `1,180.00` é americano e `1.180,00` é brasileiro, e as duas chegam em
    cotação — a segunda vem de representante nacional. Confundir uma com a
    outra erra o preço por mil vezes, então a decisão é pelo formato inteiro da
    string, nunca por trocar separador às cegas.
    """
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float, Decimal)):
        try:
            return Decimal(str(valor))
        except InvalidOperation:
            return None
    if valor is None:
        return None

    texto = str(valor).replace("\xa0", " ").strip()
    texto = _RE_MOEDA.sub("", texto).strip()
    texto = texto.replace(" ", "")
    # 'USD39.00/set' é o preço de um set, não uma divisão: a unidade é
    # decoração e o número é o mesmo com ou sem ela
    texto = _RE_UNIDADE_DE_PRECO.sub("", texto).strip()
    if not texto:
        return None
    # faixa de preço ('12.00-15.00'): fica com o maior, como o parser tabular
    if "-" in texto.strip("-") or "~" in texto:
        partes = [p for p in re.split(r"[-~]", texto) if p]
        valores = [para_preco(p) for p in partes]
        valores = [v for v in valores if v is not None]
        return max(valores) if valores else None

    if _RE_BR.match(texto):
        texto = texto.replace(".", "").replace(",", ".")
    elif _RE_US.match(texto):
        texto = texto.replace(",", "")
    elif _RE_VIRGULA_DECIMAL.match(texto):
        texto = texto.replace(",", ".")
    elif not _RE_SO_NUMERO.match(texto):
        return None

    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _e_preco_plausivel(valor: Decimal | None) -> bool:
    return valor is not None and PRECO_MINIMO <= valor <= PRECO_MAXIMO


@dataclass
class _Coluna:
    indice: int
    textos: list[str] = field(default_factory=list)
    numeros: list[Decimal] = field(default_factory=list)
    com_decimal: int = 0
    distintos: set = field(default_factory=set)
    comprimento_total: int = 0

    @property
    def preenchidas(self) -> int:
        return len(self.textos) + len(self.numeros)

    @property
    def comprimento_medio(self) -> float:
        return self.comprimento_total / len(self.textos) if self.textos else 0.0


def _grade(ws) -> list[list]:
    """A aba como lista de linhas, limitada e 0-based, com as mesclas resolvidas."""
    linhas = []
    for indice, linha in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 1, MAX_LINHAS),
            max_col=min(ws.max_column or 1, MAX_COLUNAS),
            values_only=True,
        )
    ):
        linhas.append(list(linha))
        if indice >= MAX_LINHAS:
            break
    return preencher_mesclas_verticais(linhas, ws)


def _densidade(linha) -> int:
    return sum(1 for célula in linha if texto_limpo(célula))


def _regiao_de_dados(grade) -> tuple[int, int] | None:
    """O maior trecho de linhas seguidas com pelo menos duas células cheias.

    Uma linha de uma célula só **no meio** do trecho não o interrompe: é o
    título de seção ('Refrigeration', 'Cooking') que os catálogos usam para
    separar famílias. Tratá-la como fim da tabela partia a cotação em pedaços e
    a leitura ficava com o maior deles — os produtos depois do segundo título
    sumiam sem aviso, que é o pior jeito de perder dado.

    Duas linhas esparsas seguidas continuam sendo o fim: aí já é rodapé.

    Devolve (primeira, última) em índices 0-based da grade, ou None.
    """
    melhor = atual = None
    pendente = None  # linha esparsa segurada à espera de dado depois dela
    for indice, linha in enumerate(grade):
        densidade = _densidade(linha)
        if densidade >= 2:
            if atual is None:
                atual = (indice, indice)
            else:
                atual = (atual[0], indice)
            if melhor is None or (atual[1] - atual[0]) > (melhor[1] - melhor[0]):
                melhor = atual
            pendente = None
        elif densidade == 1 and atual is not None and pendente is None:
            pendente = indice
        else:
            atual = pendente = None
    if melhor is None or (melhor[1] - melhor[0] + 1) < MINIMO_DE_LINHAS_NA_REGIAO:
        return None
    return melhor


def _cabecalho_por_vocabulario(grade, ws, primeira: int, ultima: int, limite: int = 30):
    """O cabeçalho, quando algum rótulo conhecido aparece nele.

    Devolve `(primeira, última, mapa, rótulos)` — todas as linhas do cabeçalho,
    não só a primeira, porque `Carton (mm)` mesclado sobre `L | W | H` ocupa
    duas e a de baixo não é produto.

    Vale menos que no detector — aqui basta **um** papel, porque este parser só
    roda depois de os parsers específicos terem desistido, e um `Price` isolado
    já é mais informação do que a forma sozinha.
    """
    melhor = None
    for indice in range(min(limite, len(grade))):
        # um cabeçalho começa onde há texto: deixar uma linha vazia abrir o
        # bloco faria o cabeçalho de verdade virar a *continuação* dela, e a
        # linha de títulos entraria na tabela como se fosse um produto
        if not _densidade(grade[indice]):
            continue
        fim = fim_do_cabecalho(grade, indice)
        # Um cabeçalho encosta na tabela: ou é a primeira linha dela, ou está
        # logo acima. Sem esta amarra, a linha de rodapé 'Note: prices are FOB
        # Ningbo' virava cabeçalho — ela tem a palavra que o vocabulário
        # reconhece como preço —, e a coluna 1 da tabela virava a de preço.
        if not (primeira - 1 <= fim <= ultima):
            continue
        rotulos = rotulos_achatados(grade, ws, indice, fim)
        mapa = mapear_papeis(rotulos)
        if not mapa:
            continue
        if melhor is None or len(mapa) > len(melhor[2]):
            melhor = (indice, fim, mapa, rotulos)
    return melhor


def _e_cabecalho_por_forma(grade, indice: int, ultima: int) -> bool:
    """Uma linha só de texto, com números logo abaixo, é cabeçalho.

    Quando nenhum rótulo é reconhecido, é isto que impede o cabeçalho de entrar
    na conta como se fosse produto — e de contaminar a estatística das colunas,
    que é justamente o que decide qual coluna é o quê.
    """
    linha = grade[indice]
    preenchidas = [c for c in linha if texto_limpo(c) is not None]
    if len(preenchidas) < 2:
        return False
    if any(para_preco(c) is not None for c in preenchidas):
        return False

    for abaixo in grade[indice + 1 : min(indice + 4, ultima + 1)]:
        if any(para_preco(c) is not None for c in abaixo if texto_limpo(c)):
            return True
    return False


def _colunas_da_regiao(grade, primeira: int, ultima: int) -> dict[int, _Coluna]:
    colunas: dict[int, _Coluna] = {}
    for linha in grade[primeira : ultima + 1]:
        for indice, bruto in enumerate(linha, start=1):
            texto = texto_limpo(bruto)
            if texto is None:
                continue
            coluna = colunas.setdefault(indice, _Coluna(indice=indice))
            numero = para_preco(bruto)
            if numero is not None and (
                isinstance(bruto, (int, float, Decimal)) or _RE_SO_NUMERO.match(texto)
                or _RE_BR.match(texto) or _RE_US.match(texto)
            ):
                coluna.numeros.append(numero)
                if numero != numero.to_integral_value():
                    coluna.com_decimal += 1
            else:
                coluna.textos.append(texto)
                coluna.comprimento_total += len(texto)
            coluna.distintos.add(texto)
    return colunas


def _e_sequencia(numeros: list[Decimal]) -> bool:
    """1, 2, 3, 4… é numeração de item, não preço."""
    inteiros = [n for n in numeros if n == n.to_integral_value()]
    if len(inteiros) < 3 or len(inteiros) != len(numeros):
        return False
    ordenados = sorted(set(int(n) for n in inteiros))
    return len(ordenados) >= 3 and ordenados == list(
        range(ordenados[0], ordenados[0] + len(ordenados))
    )


def _eleger_preco(colunas: dict[int, _Coluna]) -> int | None:
    melhor, melhor_ponto = None, 0.0
    for coluna in colunas.values():
        plausiveis = [n for n in coluna.numeros if _e_preco_plausivel(n)]
        if len(plausiveis) < MINIMO_PARA_ELEGER_COLUNA:
            continue
        if _e_sequencia(coluna.numeros):
            continue
        # decimal é o sinal mais forte de preço; peso e quantidade competem
        # aqui e perdem porque são inteiros na maioria das cotações
        ponto = len(plausiveis) + 2.0 * coluna.com_decimal
        if ponto > melhor_ponto:
            melhor, melhor_ponto = coluna.indice, ponto
    return melhor


_RE_COMERCIAL = re.compile(
    r"\b(fob|exw|cif|cfr|ddp|quotation|quote|cota[çc][ãa]o|proforma|price|"
    r"pre[çc]o|payment|pagamento|moq|incoterm|t/t|l/c|lead\s*time|"
    r"delivery|entrega|validity|validade)\b",
    re.IGNORECASE,
)


def _e_documento_comercial(grade) -> bool:
    """A planilha inteira parece uma oferta de fornecedor?

    Pergunta diferente de "qual coluna é o preço", e por isso respondida com
    outra evidência: a razão social no topo, ou o vocabulário de comércio
    exterior em qualquer lugar do arquivo. Um controle de estoque tem código e
    número como uma cotação tem, e não tem nada disto.
    """
    if not _identidade_fornecedor(grade).startswith("("):
        return True
    for linha in grade[:60]:
        for bruto in linha:
            texto = texto_limpo(bruto)
            if texto and _RE_COMERCIAL.search(texto):
                return True
    return False


def _parece_dinheiro(coluna: _Coluna | None) -> bool:
    """A coluna tem centavos ou traz a moeda escrita?

    Só é consultada quando o cabeçalho não disse nada e a coluna foi eleita
    preço por eliminação. Nesse ponto, "tem casas decimais" é o que resta de
    evidência de que aquilo é dinheiro e não contagem.
    """
    if coluna is None:
        return False
    return coluna.com_decimal > 0 or any(
        _RE_MOEDA.search(texto) for texto in coluna.textos
    )


def _eleger_identidade(
    colunas: dict[int, _Coluna], evitar: set
) -> tuple[int | None, float]:
    """A coluna que identifica o produto, e o quanto ela parece código.

    Três sinais, nesta ordem de força: **parecer código** (curto, sem espaço,
    com dígito), **variar a cada linha** (categoria se repete, código não) e
    **ser curta** — a coluna de descrição também varia a cada linha, e sem a
    penalidade de comprimento ela ganharia de um código de seis caracteres só
    por ter mais texto.

    A fração de código volta junto porque quem chama precisa dela para decidir
    se acredita na eleição: numa planilha sem nenhum rótulo reconhecido, é o
    único sinal que separa uma cotação de uma tabela qualquer.
    """
    melhor, melhor_ponto, melhor_codigo = None, 0.0, 0.0
    for coluna in colunas.values():
        if coluna.indice in evitar or not coluna.textos:
            continue
        curtos = [t for t in coluna.textos if len(t) <= 60]
        if len(curtos) < MINIMO_PARA_ELEGER_COLUNA:
            continue
        variedade = len(coluna.distintos) / max(len(coluna.textos), 1)
        fracao_codigo = sum(1 for t in curtos if _RE_CODIGO.match(t)) / len(curtos)
        penalidade_comprimento = 1.0 + coluna.comprimento_medio / 20.0
        ponto = (
            len(curtos) * variedade * (1.0 + 3.0 * fracao_codigo)
        ) / penalidade_comprimento
        if ponto > melhor_ponto:
            melhor, melhor_ponto, melhor_codigo = coluna.indice, ponto, fracao_codigo
    return melhor, melhor_codigo


def _identidade_do_cabecalho(rotulos: list, colunas: dict[int, _Coluna]) -> int | None:
    """A coluna de identidade entre as que o cabeçalho indica, conferida no dado.

    `Item` e `Item No.` são a mesma palavra para o vocabulário, e numa cotação
    que tem as duas a primeira é a numeração da linha (1, 2, 3) e a segunda é o
    código do produto. Escolher pela ordem entrega a numeração como identidade,
    e aí toda linha é descartada por "identidade que é só número" — a cotação
    inteira some com a mensagem de que não há produtos nela.

    A regra de desempate não é mais vocabulário: é olhar o que está **embaixo**
    do rótulo. Coluna de identidade tem texto; numeração de linha não tem.
    """
    if not rotulos:
        return None
    candidatas = mapear_todos(rotulos)
    for papel in ("modelo", "nome"):
        for indice in candidatas.get(papel, []):
            coluna = colunas.get(indice)
            if coluna is not None and coluna.textos:
                return indice
    return None


def _eleger_descricao(colunas: dict[int, _Coluna], evitar: set) -> int | None:
    melhor, maior = None, 0.0
    for coluna in colunas.values():
        if coluna.indice in evitar or len(coluna.textos) < MINIMO_PARA_ELEGER_COLUNA:
            continue
        if coluna.comprimento_medio > maior:
            melhor, maior = coluna.indice, coluna.comprimento_medio
    return melhor if maior >= 12 else None


def _embalagem_de(textos: list[str]) -> Embalagem:
    """Garimpa carton, peças/caixa e CBM de textos livres. Não avisa nada — quem
    chama sabe se uma coluna dedicada ainda vai preencher o que faltou."""
    emb = Embalagem()
    for texto in textos:
        if emb.carton_mm is None:
            dims = extrair_dimensoes_mm(texto)
            if dims:
                emb.carton_mm = dims
        if emb.pcs_por_carton is None:
            m = re.search(r"(\d+)\s*(?:pcs?|pçs?|pe[çc]as?)\s*/?\s*(?:ctn|carton|cx)", texto, re.IGNORECASE)
            if m:
                emb.pcs_por_carton = int(m.group(1))
        if emb.cbm_total is None:
            m = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:cbm|m³|m3)\b", texto, re.IGNORECASE)
            if m:
                emb.cbm_total = para_preco(m.group(1))
                emb.qty_referencia = emb.qty_referencia or 1
    return emb


def _identidade_fornecedor(grade) -> str:
    """O nome do fornecedor costuma estar nas primeiras linhas, no topo."""
    for linha in grade[:10]:
        for bruto in linha[:6]:
            texto = texto_limpo(bruto)
            if not texto or len(texto) > 80:
                continue
            if re.search(
                r"\b(ltd|limited|ltda|co\.|inc|industr|trading|technolog|"
                r"machinery|equipment|group|s\.a\.)\b",
                texto,
                re.IGNORECASE,
            ):
                return texto
    return "(fornecedor não identificado)"


def _fichas_da_aba(ws, nome_arquivo: str) -> list[Ficha]:
    grade = _grade(ws)
    if not grade:
        return []

    regiao = _regiao_de_dados(grade)
    if regiao is None:
        return []
    primeira, ultima = regiao

    achado = _cabecalho_por_vocabulario(grade, ws, primeira, ultima)
    mapa_papeis: dict[str, int] = {}
    rotulos: list = []
    if achado is not None:
        linha_cabecalho, fim_cabecalho, mapa_papeis, rotulos = achado
        # o cabeçalho não é produto: a região começa depois dele — e ele pode
        # ter mais de uma linha (rótulo mesclado em cima, sub-rótulo embaixo)
        if linha_cabecalho <= ultima and fim_cabecalho >= primeira:
            primeira = fim_cabecalho + 1
    elif _e_cabecalho_por_forma(grade, primeira, ultima):
        primeira += 1
    if primeira > ultima:
        return []

    colunas = _colunas_da_regiao(grade, primeira, ultima)
    if not colunas:
        return []

    # a caixa espalhada em três colunas numéricas, uma por dimensão
    cols_carton = mapear_todos(rotulos).get("carton_dimensao", []) if rotulos else []
    # a unidade está no rótulo de cima ('Carton (cm) L'), e sem ela o m³ sai
    # mil vezes errado
    fator_carton = 10 if any(
        re.search(r"\bcm\b", rotulos[c - 1], re.IGNORECASE)
        for c in cols_carton
        if c - 1 < len(rotulos)
    ) else 1
    # preço por faixa de quantidade: `1-49 pcs | 50-99 pcs | 100+`. Nenhuma
    # das colunas é "o" preço — todas viram variante, com a faixa por rótulo.
    faixas_de_preco = [
        (indice, faixa)
        for indice, rotulo in enumerate(rotulos, start=1)
        for faixa in (faixa_de_quantidade(rotulo),)
        if faixa
    ]

    col_preco = mapa_papeis.get("preco")
    if col_preco is None and faixas_de_preco:
        col_preco = faixas_de_preco[0][0]
    if col_preco is None:
        col_preco = _eleger_preco(colunas)
    evitar = {col_preco} if col_preco else set()
    if rotulos:
        # a conferência no dado vale mais que a ordem das colunas; se nenhuma
        # das candidatas do cabeçalho tem texto, é melhor cair na eleição por
        # forma do que aceitar uma coluna de numeração como identidade
        col_identidade = _identidade_do_cabecalho(rotulos, colunas)
    else:
        col_identidade = mapa_papeis.get("modelo") or mapa_papeis.get("nome")
    identidade_adivinhada = col_identidade is None
    fracao_codigo = 1.0
    if identidade_adivinhada:
        col_identidade, fracao_codigo = _eleger_identidade(colunas, evitar)
    if col_identidade is None:
        return []

    # Quando nem a identidade veio de um rótulo, sobram três sinais para
    # separar uma cotação de uma planilha qualquer com texto e números: haver
    # preço, a coluna de identidade parecer código de modelo, e **os números
    # parecerem dinheiro**.
    #
    # O terceiro sinal é o que separa uma cotação sem cabeçalho de um controle
    # de estoque: os dois têm código na primeira coluna e número na última, e
    # até aqui os dois passavam — o estoque virava três "produtos" com preço de
    # 14, 8 e 22 dólares. Preço de equipamento tem centavos ou traz a moeda
    # escrita; contagem de prateleira é inteiro seco. Devolver uma lista de
    # produtos inventada é pior do que devolver erro.
    #
    # O terceiro sinal admite duas provas, porque cotação de verdade nem sempre
    # tem centavos: ou os números parecem dinheiro, ou o documento inteiro
    # parece comercial (fornecedor com razão social, FOB, condição de
    # pagamento). Exigir só a primeira recusava a cotação de três linhas com
    # preços redondos, que é um arquivo que existe.
    if identidade_adivinhada and (
        col_preco is None
        or fracao_codigo < FRACAO_MINIMA_DE_CODIGO
        or not (
            _parece_dinheiro(colunas.get(col_preco)) or _e_documento_comercial(grade)
        )
    ):
        return []

    evitar.add(col_identidade)
    col_descricao = mapa_papeis.get("descricao") or _eleger_descricao(colunas, evitar)
    col_moq = mapa_papeis.get("moq")
    col_embalagem = mapa_papeis.get("embalagem")
    col_cbm = mapa_papeis.get("cbm")
    col_pcs = mapa_papeis.get("pcs_por_caixa")
    col_foto = mapa_papeis.get("foto")
    col_certificado = mapa_papeis.get("certificado")
    col_peso_liquido = mapa_papeis.get("peso_liquido")
    col_peso_bruto = mapa_papeis.get("peso_bruto")

    # Só identidade e preço contam como adivinhação: errá-las troca o produto
    # ou o número que decide a compra. A descrição eleita por comprimento, no
    # pior caso, traz texto a mais na tela — e a pessoa está olhando para ele.
    adivinhadas = [
        rotulo
        for rotulo, coluna, veio_do_mapa in (
            ("identidade do produto", col_identidade, not identidade_adivinhada),
            ("preço", col_preco, "preco" in mapa_papeis or bool(faixas_de_preco)),
        )
        if coluna is not None and not veio_do_mapa
    ]
    # com identidade e preço vindos dos títulos, a leitura é tão boa quanto a
    # de um parser específico; dizer "não reconhecido" aí seria alarme falso, e
    # alarme falso repetido é o que faz as pessoas pararem de ler os avisos
    confianca = "baixa" if adivinhadas else "media"

    fotos = fotos_por_ancora_xlsx(ws)
    fichas: list[Ficha] = []
    categoria_corrente: str | None = None

    for indice_linha in range(primeira, ultima + 1):
        linha = grade[indice_linha]
        numero_excel = indice_linha + 1

        def celula(coluna):
            if not coluna or coluna > len(linha):
                return None
            return linha[coluna - 1]

        identidade = texto_limpo(celula(col_identidade))
        if not identidade or _RE_RODAPE.match(identidade):
            continue
        # uma célula que é só número não identifica produto nenhum
        if _RE_SO_NUMERO.match(identidade):
            continue

        preco_valor = para_preco(celula(col_preco)) if col_preco else None
        if preco_valor is not None and not _e_preco_plausivel(preco_valor):
            preco_valor = None
        descricao = texto_limpo(celula(col_descricao)) or ""

        # Linha com a identidade preenchida e mais nada é título de seção
        # ('Refrigeration'), não produto. Ela não se perde: vira a categoria
        # dos produtos que vêm abaixo dela, que é o que ela significa para
        # quem lê a cotação.
        if preco_valor is None and not descricao:
            if _densidade(linha) == 1 and not _RE_CODIGO.match(identidade):
                categoria_corrente = identidade
            continue

        # Duas linhas seguidas com o mesmo código são o mesmo produto: é o
        # bloco de especificação que o fornecedor escreve em várias linhas,
        # com o código mesclado por cima delas. Cada linha vira uma linha da
        # descrição, e o que a primeira não trouxe (preço, embalagem) as
        # seguintes completam — nunca sobrescrevem.
        if fichas and fichas[-1].modelo == identidade:
            anterior = fichas[-1]
            if descricao and descricao not in anterior.descricao_bruta.splitlines():
                anterior.descricao_bruta = (
                    f"{anterior.descricao_bruta}\n{descricao}"
                    if anterior.descricao_bruta
                    else descricao
                )
            if preco_valor is not None and not anterior.precos:
                anterior.precos.append(
                    Preco(
                        valor=preco_valor,
                        moeda="USD",
                        incoterm=None,
                        rotulo="padrão",
                        moq=primeiro_inteiro(texto_limpo(celula(col_moq)))
                        if col_moq
                        else None,
                        origem=anterior.origem,
                    )
                )
            continue

        origem = Origem(
            arquivo=nome_arquivo,
            aba_ou_pagina=ws.title,
            celula_ou_bbox=f"linha {numero_excel}",
            confianca=confianca,
        )

        moq = primeiro_inteiro(texto_limpo(celula(col_moq))) if col_moq else None
        precos = []
        if preco_valor is not None:
            precos.append(
                Preco(
                    valor=preco_valor,
                    moeda="USD",
                    incoterm=None,
                    rotulo=(faixas_de_preco[0][1] if faixas_de_preco else "padrão"),
                    moq=moq,
                    origem=origem,
                )
            )
        for coluna_faixa, faixa in faixas_de_preco[1:]:
            valor_faixa = para_preco(celula(coluna_faixa))
            if valor_faixa is None or not _e_preco_plausivel(valor_faixa):
                continue
            precos.append(
                Preco(
                    valor=valor_faixa,
                    moeda="USD",
                    incoterm=None,
                    rotulo=faixa,
                    moq=moq,
                    origem=origem,
                )
            )

        textos_embalagem = [
            t
            for t in (
                texto_limpo(celula(col_embalagem)),
                texto_limpo(celula(col_cbm)),
                descricao,
            )
            if t
        ]
        embalagem = _embalagem_de(textos_embalagem)
        # a coluna dedicada vence o que foi garimpado do texto: 'PCS/CTN' é
        # declaração do fornecedor, o regex na descrição é interpretação nossa
        if col_pcs:
            pcs = primeiro_inteiro(texto_limpo(celula(col_pcs)))
            if pcs:
                embalagem.pcs_por_carton = pcs
        if col_cbm and embalagem.cbm_total is None:
            cbm = para_preco(celula(col_cbm))
            if cbm is not None and cbm > 0:
                embalagem.cbm_total = cbm
                embalagem.qty_referencia = embalagem.pcs_por_carton or 1
        # a caixa em três colunas: `Carton (mm)` sobre `L | W | H`
        if embalagem.carton_mm is None and len(cols_carton) == 3:
            medidas = [para_preco(celula(c)) for c in cols_carton]
            if all(m is not None and m > 0 for m in medidas):
                embalagem.carton_mm = tuple(
                    int(round(float(m) * fator_carton)) for m in medidas
                )
        for coluna_peso, atributo in (
            (col_peso_liquido, "peso_liquido_kg"),
            (col_peso_bruto, "peso_bruto_kg"),
        ):
            if not coluna_peso:
                continue
            peso = para_preco(celula(coluna_peso))
            if peso is not None and peso > 0:
                setattr(embalagem, atributo, peso)

        avisos_emb = []
        if embalagem.carton_mm and embalagem.pcs_por_carton is None and embalagem.cbm_total is None:
            avisos_emb.append(
                "medida de caixa encontrada mas peças por caixa não — m³ "
                "unitário não calculável"
            )

        avisos = []
        if adivinhadas:
            avisos.append(
                "layout de cotação não reconhecido — a ferramenta leu esta "
                "planilha pela forma da tabela, e adivinhou a coluna de "
                + ", ".join(adivinhadas)
                + ". CONFIRA cada campo antes de gravar."
            )
        else:
            avisos.append(
                "cotação em layout novo, lida pelos títulos das colunas — "
                "confira o preço e a embalagem antes de gravar"
            )
        if not precos:
            avisos.append("preço não informado na cotação")
        avisos.extend(avisos_emb)

        foto = foto_formato = None
        if col_foto:
            achada = fotos.get((numero_excel, col_foto))
            if achada:
                foto, foto_formato = achada
        if foto is None:
            for (linha_ancora, _), (dados, formato) in sorted(fotos.items()):
                if linha_ancora == numero_excel:
                    foto, foto_formato = dados, formato
                    break

        certificacoes = []
        if col_certificado:
            cert = texto_limpo(celula(col_certificado))
            if cert:
                certificacoes = [cert]

        fichas.append(
            Ficha(
                fornecedor=_identidade_fornecedor(grade),
                contato=None,
                data_cotacao=None,
                validade=None,
                modelo=identidade,
                descricao_bruta=descricao,
                categoria=categoria_corrente,
                specs={},
                precos=precos,
                embalagem=embalagem,
                certificacoes=certificacoes,
                foto=foto,
                foto_formato=foto_formato,
                origem=origem,
                avisos=avisos,
            )
        )

    return fichas


def parse_xlsx_generico(caminho: Path) -> list[Ficha]:
    """Lê uma planilha de layout desconhecido. Devolve [] quando não há tabela."""
    caminho = Path(caminho)
    wb = openpyxl.load_workbook(caminho, data_only=True)
    try:
        wb_imgs = openpyxl.load_workbook(caminho)
    except Exception:
        wb_imgs = None

    fichas: list[Ficha] = []
    try:
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            if ws.sheet_state != "visible":
                continue
            if wb_imgs is not None and nome_aba in wb_imgs.sheetnames:
                # valores vêm de `ws` (data_only), imagens do workbook sem cache
                ws._images = getattr(wb_imgs[nome_aba], "_images", [])
            fichas.extend(_fichas_da_aba(ws, caminho.name))
            if len(fichas) >= MAXIMO_DE_FICHAS:
                fichas = fichas[:MAXIMO_DE_FICHAS]
                for ficha in fichas:
                    ficha.avisos.append(
                        f"a leitura parou em {MAXIMO_DE_FICHAS} produtos — se a "
                        "cotação tem mais, ela precisa ser lançada à mão"
                    )
                break
    finally:
        wb.close()
        if wb_imgs is not None:
            wb_imgs.close()

    return fichas
