"""Parâmetros do motor de custo — lidos da aba `Pesos`, nunca embutidos em fórmula.

PLANO.md seção 6.5. A aba `Pesos` já é o painel de parâmetros da planilha, com
a convenção coluna C = rótulo, D = valor, E = nota. Os parâmetros de custo
seguem a mesma convenção, numa seção nova abaixo da existente.

Cada parâmetro carrega um selo de **confirmado ou não**. Um número que veio de
resposta do dono do negócio vale diferente de um default legal que ninguém
conferiu ainda, e essa diferença precisa chegar ao relatório — senão o
`custo estimado` da coluna O parece igualmente sólido nos dois casos.

Respostas já obtidas (PLANO.md seção 13):
  13.1  regime = Lucro Presumido  -> PIS e COFINS da importação **não** geram
        crédito; são custo. É a resposta que mais muda o número final.
  13.2  ICMS da importação é integralmente creditado.
  13.3  contêiner ~5.000 USD, 70% de aproveitamento do volume nominal.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from typing import Iterable

REGIME_PRESUMIDO = "Lucro Presumido"
REGIME_REAL = "Lucro Real"

# um 40HQ tem cerca de 76 m³ nominais (PLANO.md 6.3)
M3_NOMINAIS_40HQ = Decimal("76")


@dataclass
class Parametro:
    chave: str
    rotulo: str
    valor: Decimal | str
    unidade: str = ""
    nota: str = ""
    confirmado: bool = False
    # linha em que o parâmetro está na aba `Pesos`, quando veio de lá. É o que
    # permite a coluna P do Funil referenciar `Pesos!$D$n` em vez de trazer o
    # markup embutido na fórmula (PLANO.md 3.5.7).
    linha: int | None = None

    @property
    def decimal(self) -> Decimal:
        if isinstance(self.valor, Decimal):
            return self.valor
        return Decimal(str(self.valor))


def _p(chave, rotulo, valor, unidade="", nota="", confirmado=False) -> Parametro:
    valor_final = valor if isinstance(valor, str) else Decimal(str(valor))
    return Parametro(chave, rotulo, valor_final, unidade, nota, confirmado)


A_CONFERIR = "A conferir com o despachante antes de usar em decisão."


def parametros_padrao() -> list[Parametro]:
    """Defaults. Os confirmados vêm das respostas da seção 13; os demais são
    alíquotas legais gerais ou estimativas que ainda precisam de conferência."""
    return [
        _p("cambio_usd_brl", "Câmbio USD/BRL", "5.2", "BRL/USD",
           "Espelha Pesos!D24. Trocar quando o câmbio de fechamento mudar.", True),
        _p("regime_tributario", "Regime tributário", REGIME_PRESUMIDO, "",
           "Resposta 13.1. No Presumido (cumulativo) PIS e COFINS não geram crédito.", True),
        _p("ano_base", "Ano-base do cálculo", "2026", "",
           "O motor é função do ano por causa da reforma tributária (PLANO 6.7).", True),
        _p("icms_integralmente_creditado", "ICMS da importação é integralmente creditado", "1", "sim=1",
           "Resposta 13.2. Se surgir ST ou benefício fiscal, mudar para 0.", True),

        _p("custo_conteiner_usd", "Custo do contêiner 40HQ", "5000", "USD",
           "Resposta 13.3. Varia muito; 5.000 USD é a premissa de planejamento.", True),
        _p("aproveitamento_conteiner", "Aproveitamento do contêiner", "0.70", "fração",
           "Resposta 13.3. 70% do volume nominal.", True),
        _p("m3_nominais_conteiner", "Volume nominal do contêiner 40HQ", str(M3_NOMINAIS_40HQ), "m³",
           "Valor de catálogo do 40HQ.", True),

        _p("seguro_percentual", "Seguro internacional", "0.003", "fração sobre FOB+frete",
           A_CONFERIR),
        _p("aliquota_afrmm", "Alíquota AFRMM sobre o frete marítimo", "0.08", "fração",
           "8% na navegação de longo curso (Lei 14.301/2022). " + A_CONFERIR),
        _p("taxa_siscomex_di", "Taxa Siscomex por DI", "154.23", "BRL",
           A_CONFERIR),
        _p("unidades_por_importacao", "Unidades por importação (rateio dos custos fixos)",
           "500", "unidades",
           "Divisor do Siscomex e do desembaraço, que são por DI e não por peça. "
           "O MOQ da cotação é a melhor estimativa quando existe. " + A_CONFERIR),
        _p("despesas_desembaraco_di", "Despesas de desembaraço por DI", "0", "BRL",
           "THC, capatazia, armazenagem, honorários. " + A_CONFERIR),
        _p("frete_interno_por_m3", "Frete interno porto→empresa", "0", "BRL/m³",
           A_CONFERIR),

        _p("aliquota_pis_importacao", "Alíquota PIS-Importação", "0.021", "fração",
           "Alíquota geral da Lei 10.865/2004. " + A_CONFERIR),
        _p("aliquota_cofins_importacao", "Alíquota COFINS-Importação", "0.0965", "fração",
           "Alíquota geral da Lei 10.865/2004. " + A_CONFERIR),
        _p("aliquota_icms_importacao", "Alíquota ICMS de importação", "0.18", "fração",
           "Depende do estado e do regime. " + A_CONFERIR),

        _p("markup_minimo_revenda", "Markup mínimo de revenda", "2.2", "múltiplo",
           "Substitui o multiplicador solto da coluna P do Funil. " + A_CONFERIR),
        _p("markup_varejo", "Markup varejo", "1.5", "múltiplo", A_CONFERIR),
    ]


@dataclass
class ParametrosCusto:
    itens: dict[str, Parametro] = field(default_factory=dict)

    @classmethod
    def padrao(cls) -> "ParametrosCusto":
        return cls({p.chave: p for p in parametros_padrao()})

    def __getitem__(self, chave: str) -> Parametro:
        if chave not in self.itens:
            raise KeyError(f"parâmetro de custo desconhecido: {chave}")
        return self.itens[chave]

    def valor(self, chave: str) -> Decimal:
        return self[chave].decimal

    def texto(self, chave: str) -> str:
        return str(self[chave].valor)

    def definir(
        self,
        chave: str,
        valor,
        confirmado: bool | None = None,
        linha: int | None = None,
    ) -> None:
        atual = self[chave]
        atual.valor = valor if isinstance(valor, str) else Decimal(str(valor))
        if confirmado is not None:
            atual.confirmado = confirmado
        if linha is not None:
            atual.linha = linha

    @property
    def nao_confirmados(self) -> list[Parametro]:
        return [p for p in self.itens.values() if not p.confirmado]

    @property
    def m3_uteis_conteiner(self) -> Decimal:
        return self.valor("m3_nominais_conteiner") * self.valor("aproveitamento_conteiner")

    def premissas(self, chaves: Iterable[str]) -> dict[str, str]:
        """Snapshot legível para a memória de cálculo."""
        saida = {}
        for chave in chaves:
            p = self[chave]
            selo = "" if p.confirmado else "  [A CONFERIR]"
            unidade = f" {p.unidade}" if p.unidade else ""
            saida[p.rotulo] = f"{p.valor}{unidade}{selo}"
        return saida


# --------------------------------------------------------- leitura da planilha

LINHA_INICIAL_SECAO = 44
COLUNA_ROTULO = 3  # C
COLUNA_VALOR = 4  # D
COLUNA_NOTA = 5  # E
COLUNA_CONFIRMADO = 6  # F
TITULO_SECAO = "Parâmetros de custo de importação"
CABECALHOS_SECAO = ("Parâmetro", "Valor", "Observação", "Confirmado?")

_SIM = frozenset({"sim", "s", "1", "true", "verdadeiro", "x"})


def _ler_confirmado(celula, nota: str) -> bool:
    """O selo mora numa coluna própria justamente para ser editável por quem
    confirmou o número. Sem ela — planilha antiga, coluna apagada — cai no
    texto da observação, que é onde o aviso 'a conferir' está escrito."""
    if celula is not None and str(celula).strip():
        return str(celula).strip().lower() in _SIM
    return A_CONFERIR.split(".")[0] not in nota


def ler_da_planilha(caminho, aba: str = "Pesos") -> ParametrosCusto:
    """Lê os parâmetros gravados na aba `Pesos`; o que não estiver lá fica no default.

    Também reaproveita `Pesos!D24` como câmbio, que é onde ele já mora
    (PLANO.md 3.2) — a ferramenta lê daqui, nunca embute número em fórmula.
    """
    import openpyxl

    parametros = ParametrosCusto.padrao()
    wb = openpyxl.load_workbook(caminho, data_only=True)
    try:
        if aba not in wb.sheetnames:
            return parametros
        ws = wb[aba]

        cambio = ws.cell(row=24, column=COLUNA_VALOR).value
        if isinstance(cambio, (int, float)) and cambio > 0:
            parametros.definir("cambio_usd_brl", Decimal(str(cambio)), confirmado=True)

        rotulo_para_chave = {p.rotulo: p.chave for p in parametros.itens.values()}
        for linha in range(LINHA_INICIAL_SECAO, ws.max_row + 1):
            rotulo = ws.cell(row=linha, column=COLUNA_ROTULO).value
            valor = ws.cell(row=linha, column=COLUNA_VALOR).value
            if rotulo is None or valor is None:
                continue
            chave = rotulo_para_chave.get(str(rotulo).strip())
            if chave is None:
                continue
            nota = str(ws.cell(row=linha, column=COLUNA_NOTA).value or "")
            confirmado = _ler_confirmado(
                ws.cell(row=linha, column=COLUNA_CONFIRMADO).value, nota
            )
            parametros.definir(chave, valor, confirmado=confirmado, linha=linha)
    finally:
        wb.close()

    return parametros
