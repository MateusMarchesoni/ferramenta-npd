"""Tabela NCM → alíquotas de II e IPI.

PLANO.md seção 6.6: **nenhuma cotação traz o NCM.** É classificação fiscal
brasileira, definida por despachante, não pelo fornecedor chinês. A resposta
13.4 confirma que vem de consulta ao despachante antes da importação.

Portanto:

- o NCM é campo de entrada humana no momento da seleção do produto;
- a ferramenta pode **sugerir** o NCM de um produto já cadastrado da mesma
  categoria, sempre marcado como sugestão a confirmar;
- **sem NCM, o custo econômico não é calculado** — a linha entra na planilha
  com FOB e m³ preenchidos e a coluna de custo vazia.

Nunca inferir NCM por semelhança de descrição sem marcação explícita:
classificação errada gera multa, não só número errado.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal

_RE_SO_DIGITOS = re.compile(r"\D")


def normalizar_ncm(ncm: str | int | None) -> str | None:
    """'8516.60.00' e '85166000' viram '85166000'."""
    if ncm is None:
        return None
    digitos = _RE_SO_DIGITOS.sub("", str(ncm))
    return digitos if len(digitos) == 8 else None


def formatar_ncm(ncm: str) -> str:
    return f"{ncm[:4]}.{ncm[4:6]}.{ncm[6:8]}" if len(ncm) == 8 else ncm


def _como_data(valor) -> date | None:
    """A data pode voltar da planilha como data de verdade ou como o texto ISO
    que a ferramenta grava. Qualquer outra coisa é ausência de conferência."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    try:
        return date.fromisoformat(str(valor).strip()[:10])
    except (ValueError, TypeError):
        return None


@dataclass
class AliquotasNCM:
    ncm: str
    descricao: str
    aliquota_ii: Decimal
    aliquota_ipi: Decimal
    observacao: str = ""
    data_conferencia: date | None = None
    responsavel: str = ""

    @property
    def conferida(self) -> bool:
        return self.data_conferencia is not None and bool(self.responsavel)


@dataclass
class TabelaNCM:
    entradas: dict[str, AliquotasNCM] = field(default_factory=dict)

    def adicionar(self, aliquotas: AliquotasNCM) -> None:
        chave = normalizar_ncm(aliquotas.ncm)
        if chave is None:
            raise ValueError(f"NCM inválido: {aliquotas.ncm!r} — precisa ter 8 dígitos")
        aliquotas.ncm = chave
        self.entradas[chave] = aliquotas

    def buscar(self, ncm: str | int | None) -> AliquotasNCM | None:
        chave = normalizar_ncm(ncm)
        return self.entradas.get(chave) if chave else None

    def __len__(self) -> int:
        return len(self.entradas)

    @classmethod
    def de_linhas(cls, linhas) -> "TabelaNCM":
        """linhas: iterável de (ncm, descricao, ii, ipi, obs, data, responsavel)."""
        tabela = cls()
        for linha in linhas:
            campos = list(linha) + [None] * (7 - len(linha))
            ncm, descricao, ii, ipi, obs, data_conf, resp = campos[:7]
            if normalizar_ncm(ncm) is None:
                continue
            tabela.adicionar(
                AliquotasNCM(
                    ncm=str(ncm),
                    descricao=str(descricao or ""),
                    aliquota_ii=Decimal(str(ii or 0)),
                    aliquota_ipi=Decimal(str(ipi or 0)),
                    observacao=str(obs or ""),
                    data_conferencia=_como_data(data_conf),
                    responsavel=str(resp or ""),
                )
            )
        return tabela


TITULO_TABELA = "Tabela NCM → alíquotas"
CABECALHOS_TABELA = (
    "NCM",
    "Descrição",
    "Alíquota II",
    "Alíquota IPI",
    "Observação",
    "Data da última conferência",
    "Responsável pela conferência",
)


def ler_da_planilha(caminho, aba: str = "Pesos") -> TabelaNCM:
    """Lê a tabela NCM gravada na aba `Pesos`, abaixo dos parâmetros de custo."""
    import openpyxl

    wb = openpyxl.load_workbook(caminho, data_only=True)
    try:
        if aba not in wb.sheetnames:
            return TabelaNCM()
        ws = wb[aba]

        linha_cabecalho = None
        for linha in range(1, ws.max_row + 1):
            if str(ws.cell(row=linha, column=3).value or "").strip() == CABECALHOS_TABELA[0]:
                linha_cabecalho = linha
                break
        if linha_cabecalho is None:
            return TabelaNCM()

        linhas = []
        for linha in range(linha_cabecalho + 1, ws.max_row + 1):
            valores = [ws.cell(row=linha, column=3 + i).value for i in range(7)]
            if valores[0] is None:
                continue
            linhas.append(valores)
        return TabelaNCM.de_linhas(linhas)
    finally:
        wb.close()


def sugerir_por_categoria(
    tabela: TabelaNCM, categoria: str | None
) -> tuple[AliquotasNCM | None, str | None]:
    """Sugere um NCM já cadastrado cuja descrição case com a categoria.

    Devolve sempre um aviso junto: a sugestão nunca pode ser usada sem
    confirmação humana (PLANO.md 6.6).
    """
    if not categoria:
        return None, None
    alvo = {p for p in re.findall(r"\w+", categoria.lower()) if len(p) > 3}
    if not alvo:
        return None, None

    melhor = None
    melhor_pontos = 0
    for entrada in tabela.entradas.values():
        palavras = {p for p in re.findall(r"\w+", entrada.descricao.lower()) if len(p) > 3}
        pontos = len(alvo & palavras)
        if pontos > melhor_pontos:
            melhor, melhor_pontos = entrada, pontos

    if melhor is None:
        return None, None
    return melhor, (
        f"NCM {formatar_ncm(melhor.ncm)} é apenas uma SUGESTÃO por semelhança de "
        f"categoria ('{melhor.descricao}') — confirmar com o despachante antes de "
        "usar. Classificação errada gera multa."
    )
