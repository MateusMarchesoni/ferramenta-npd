"""Transforma cada caso do catálogo num .xlsx de verdade.

Os arquivos são gerados numa pasta temporária a cada execução, e não guardados
no repositório: um .xlsx commitado vira, com o tempo, mais uma amostra
conhecida — exatamente o problema que este corpus existe para evitar. A fonte
da verdade é `casos.py`, que aparece no diff.

Para olhar um caso com o Excel aberto:

    python -m tests.corpus.gerar --pasta /tmp/corpus
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

from tests.corpus.casos import ABA_EXTRA, CASOS, RECUSAS, Caso, CasoRecusa


def _escrever(ws, linhas) -> None:
    for indice, linha in enumerate(linhas, start=1):
        for coluna, valor in enumerate(linha, start=1):
            if valor is not None:
                ws.cell(row=indice, column=coluna, value=valor)


def gerar_caso(caso: Caso, pasta: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = caso.aba
    _escrever(ws, caso.linhas)
    for intervalo in caso.mesclas:
        ws.merge_cells(intervalo)

    extra = ABA_EXTRA.get(caso.nome)
    if extra is not None:
        titulo, linhas = extra
        _escrever(wb.create_sheet(titulo), linhas)

    caminho = pasta / f"{caso.nome}.xlsx"
    wb.save(caminho)
    wb.close()
    return caminho


def gerar_recusa(caso: CasoRecusa, pasta: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = caso.aba
    _escrever(ws, caso.linhas)
    caminho = pasta / f"{caso.nome}.xlsx"
    wb.save(caminho)
    wb.close()
    return caminho


def gerar_tudo(pasta: Path) -> dict[str, Path]:
    """Gera todos os casos (inclusive os de recusa) e devolve {nome: caminho}."""
    pasta = Path(pasta)
    pasta.mkdir(parents=True, exist_ok=True)
    caminhos = {caso.nome: gerar_caso(caso, pasta) for caso in CASOS}
    caminhos.update({caso.nome: gerar_recusa(caso, pasta) for caso in RECUSAS})
    return caminhos


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pasta", type=Path, default=Path("saida/corpus"))
    args = parser.parse_args()
    caminhos = gerar_tudo(args.pasta)
    print(f"{len(caminhos)} arquivos em {args.pasta.resolve()}")
    for nome in sorted(caminhos):
        print(f"  {nome}.xlsx")


if __name__ == "__main__":
    main()
