"""Critério de aceite da Etapa 7 (PLANO.md seção 11).

    "uma pessoa que não é você consegue, sem instrução verbal, abrir uma
     cotação, escolher dois produtos, informar o NCM e gravar."

A interface é a própria planilha (resposta 13.5): uma aba `Candidatos` com
caixa de seleção. O teste percorre exatamente esse caminho — abrir, marcar
dois, preencher NCM, conferir, gravar — e verifica o que a pessoa veria.

"Sem instrução verbal" é testável em parte: as instruções têm que estar na
própria aba, e as escolhas não podem sumir entre um passo e outro. O resto
depende de alguém de verdade sentar na frente, e isso está no roteiro de
`tests/manual_etapa7.py`.
"""
from __future__ import annotations

import re
import shutil
import zipfile
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from npd_tool.cli import main
from npd_tool.custo import ncm as mod_ncm
from npd_tool.custo import parametros as mod_par
from npd_tool.custo.ncm import AliquotasNCM, TabelaNCM
from npd_tool.escrita.ooxml import NomeRepetido, escrever_parametros_custo, inserir_produtos
from npd_tool.ui import candidatos as mod_candidatos
from npd_tool.ui.candidatos import (
    COL_MARCAR,
    COL_MARCA,
    COL_NCM,
    COL_NOME,
    NOME_ABA,
    PRIMEIRA_LINHA,
    SelecaoInvalida,
)

FIXTURES = Path(__file__).parent / "fixtures"
NPD = FIXTURES / "NPD_2026_04_08_26.xlsx"
FRESPRO = FIXTURES / "Convection Oven project Quotation from Frespro--20260713.xlsx"
ASTAR = FIXTURES / "Astar~Milton Quotation.pdf"


# --------------------------------------------------------------- ferramentas
#
# O gestor digita no Excel, que preserva as fotos dentro das células. Um teste
# não pode usar openpyxl para simular isso — ele salvaria o arquivo sem as
# partes de rich value e apagaria as 71 fotos. Então a edição é feita no XML.

def digitar_na_aba(caminho: Path, edicoes: dict[str, str]) -> None:
    with zipfile.ZipFile(caminho) as z:
        partes = {i.filename: z.read(i.filename) for i in z.infolist()}
        infos = {i.filename: i for i in z.infolist()}
        ordem = [i.filename for i in z.infolist()]

    workbook = partes["xl/workbook.xml"].decode("utf-8")
    rid = re.search(rf'<sheet name="{NOME_ABA}"[^>]*r:id="(rId\d+)"', workbook).group(1)
    rels = partes["xl/_rels/workbook.xml.rels"].decode("utf-8")
    alvo = re.search(rf'<Relationship Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1)
    parte = "xl/" + alvo

    xml = partes[parte].decode("utf-8")
    for referencia, valor in edicoes.items():
        padrao = re.compile(rf'<c r="{referencia}"[^>]*?(?:/>|>.*?</c>)', re.DOTALL)
        assert padrao.search(xml), f"célula {referencia} não existe na aba"
        xml = padrao.sub(
            f'<c r="{referencia}" t="inlineStr"><is><t>{valor}</t></is></c>', xml, count=1
        )
    partes[parte] = xml.encode("utf-8")

    with zipfile.ZipFile(caminho, "w", zipfile.ZIP_DEFLATED) as z:
        for nome in ordem:
            z.writestr(infos[nome], partes[nome])


@pytest.fixture
def planilha(tmp_path):
    """Uma NPD com a Etapa 5 já feita — parâmetros e NCM na aba `Pesos`."""
    destino = tmp_path / "NPD.xlsx"
    tabela = TabelaNCM()
    tabela.adicionar(
        AliquotasNCM("85166000", "Fornos elétricos", Decimal("0.20"), Decimal("0.10"))
    )
    escrever_parametros_custo(
        NPD, destino, mod_par.ParametrosCusto.padrao(), tabela
    )
    return destino


def _celula(linha: int, coluna: str) -> str:
    return f"{coluna}{linha}"


# ---------------------------------------------- o caminho inteiro, uma vez só

def test_o_caminho_do_aceite_do_comeco_ao_fim(planilha, capsys):
    """Abrir uma cotação, escolher dois produtos, informar o NCM e gravar."""
    assert main(["--npd", str(planilha), "abrir", str(FRESPRO)]) == 0

    wb = openpyxl.load_workbook(planilha)
    try:
        aba = wb[NOME_ABA]
        modelos = [
            aba[f"D{linha}"].value
            for linha in range(PRIMEIRA_LINHA, aba.max_row + 1)
            if aba[f"D{linha}"].value
        ]
    finally:
        wb.close()
    assert modelos == ["FD-52A", "FD-65G"]

    primeira, segunda = PRIMEIRA_LINHA, PRIMEIRA_LINHA + 1
    digitar_na_aba(
        planilha,
        {
            _celula(primeira, COL_MARCAR): "x",
            _celula(primeira, COL_NCM): "8516.60.00",
            _celula(primeira, COL_MARCA): "Marchesoni",
            _celula(segunda, COL_MARCAR): "x",
            _celula(segunda, COL_NCM): "85166000",
            _celula(segunda, COL_MARCA): "MarcPro",
        },
    )

    assert main(["--npd", str(planilha), "conferir"]) == 0
    assert main(["--npd", str(planilha), "gravar"]) == 0

    wb = openpyxl.load_workbook(planilha)
    try:
        funil, prio = wb["Funil"], wb["Priorizacao"]
        assert funil["E91"].value == "Commercial convection oven FD-52A"
        assert funil["E92"].value == "Commercial convection oven FD-65G"
        assert prio["A90"].value == funil["E91"].value
        assert prio["A91"].value == funil["E92"].value
        assert funil["F91"].value == "Marchesoni"
        assert funil["F92"].value == "MarcPro"
        assert funil["M91"].value == 156.2  # regra do maior valor, com bandeja
    finally:
        wb.close()

    assert (planilha.parent / "relatorios").is_dir()
    assert list((planilha.parent / "relatorios").glob("*.md"))


# ------------------------------------------------- a aba se explica sozinha

def test_a_aba_traz_as_instrucoes_e_os_cabecalhos(planilha):
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    wb = openpyxl.load_workbook(planilha)
    try:
        aba = wb[NOME_ABA]
        instrucoes = " ".join(
            str(aba[f"A{linha}"].value or "")
            for linha in range(1, mod_candidatos.LINHA_CABECALHO)
        )
        assert "COMO USAR" in instrucoes
        assert "Marque com x" in instrucoes
        assert "NCM" in instrucoes
        # a aba precisa dizer o comando literal: "rode CONFERIR" não basta para
        # quem nunca abriu um terminal
        assert "npd-tool conferir" in instrucoes
        assert "npd-tool gravar" in instrucoes
        assert "backup" in instrucoes

        cabecalhos = [
            aba.cell(row=mod_candidatos.LINHA_CABECALHO, column=coluna).value
            for coluna in range(1, len(mod_candidatos.CABECALHOS) + 1)
        ]
        assert cabecalhos[0] == "Marcar"
        assert "NCM" in cabecalhos[9]
    finally:
        wb.close()


def test_a_aba_mostra_o_que_o_gestor_usa_para_reconhecer_o_produto(planilha):
    """Sem foto a seleção fica lenta: o gestor reconhece o produto pela imagem,
    não pelo código (PLANO.md seção 8)."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    wb = openpyxl.load_workbook(planilha)
    try:
        aba = wb[NOME_ABA]
        linha = PRIMEIRA_LINHA
        assert aba[f"B{linha}"].value == "#VALUE!"  # é assim que a foto aparece
        assert aba[f"C{linha}"].value  # fornecedor
        assert aba[f"D{linha}"].value == "FD-52A"
        assert aba[f"E{linha}"].value  # nome sugerido, editável
        assert aba[f"F{linha}"].value  # descrição
        assert aba[f"G{linha}"].value == 156.2  # preço escolhido
    finally:
        wb.close()


def test_marca_tem_lista_suspensa(planilha):
    """Digitar 'marchesoni' minúsculo faria a marca não casar."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    with zipfile.ZipFile(planilha) as z:
        workbook = z.read("xl/workbook.xml").decode("utf-8")
        rid = re.search(rf'<sheet name="{NOME_ABA}"[^>]*r:id="(rId\d+)"', workbook).group(1)
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        alvo = re.search(rf'<Relationship Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1)
        aba_xml = z.read("xl/" + alvo).decode("utf-8")

    assert "<dataValidations" in aba_xml
    assert "Marchesoni,MarcPro" in aba_xml


# ------------------------------------------- o que o gestor digitou não some

def test_conferir_nao_apaga_o_que_foi_marcado(planilha):
    """A aba é reescrita para receber a prévia do custo. Se a reescrita levasse
    junto as marcações, a pessoa marcaria dez produtos, pediria a prévia e
    voltaria para uma aba em branco."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    linha = PRIMEIRA_LINHA
    digitar_na_aba(
        planilha,
        {
            _celula(linha, COL_MARCAR): "x",
            _celula(linha, COL_NCM): "85166000",
            _celula(linha, COL_MARCA): "Marchesoni",
            _celula(linha, COL_NOME): "Forno de convecção 52L",
        },
    )
    main(["--npd", str(planilha), "conferir"])

    wb = openpyxl.load_workbook(planilha)
    try:
        aba = wb[NOME_ABA]
        assert aba[f"{COL_MARCAR}{linha}"].value == "x"
        assert aba[f"{COL_NCM}{linha}"].value == "85166000"
        assert aba[f"{COL_MARCA}{linha}"].value == "Marchesoni"
        assert aba[f"{COL_NOME}{linha}"].value == "Forno de convecção 52L"
        assert aba[f"{mod_candidatos.COL_CUSTO}{linha}"].value  # a prévia chegou
    finally:
        wb.close()


def test_a_ordem_das_linhas_nao_muda_entre_os_passos(planilha):
    """Marcar a linha 8 e encontrar outro produto nela depois destrói a
    confiança na ferramenta mais rápido que qualquer bug de conta."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO), str(ASTAR)])

    def modelos():
        wb = openpyxl.load_workbook(planilha)
        try:
            aba = wb[NOME_ABA]
            return [
                aba[f"D{linha}"].value
                for linha in range(PRIMEIRA_LINHA, aba.max_row + 1)
                if aba[f"D{linha}"].value
            ]
        finally:
            wb.close()

    antes = modelos()
    digitar_na_aba(
        planilha,
        {
            _celula(PRIMEIRA_LINHA, COL_MARCAR): "x",
            _celula(PRIMEIRA_LINHA, COL_NCM): "85166000",
            _celula(PRIMEIRA_LINHA, COL_MARCA): "Marchesoni",
        },
    )
    main(["--npd", str(planilha), "conferir"])
    assert modelos() == antes


def test_gravar_desmarca_o_que_entrou(planilha):
    """Rodar `gravar` duas vezes por engano inseriria o mesmo produto de novo."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    linha = PRIMEIRA_LINHA
    digitar_na_aba(
        planilha,
        {
            _celula(linha, COL_MARCAR): "x",
            _celula(linha, COL_NCM): "85166000",
            _celula(linha, COL_MARCA): "Marchesoni",
        },
    )
    assert main(["--npd", str(planilha), "gravar"]) == 0

    wb = openpyxl.load_workbook(planilha)
    try:
        aba = wb[NOME_ABA]
        assert aba[f"{COL_MARCAR}{linha}"].value in (None, "")
        assert "Funil linha 91" in str(aba[f"{mod_candidatos.COL_CUSTO}{linha}"].value)
    finally:
        wb.close()

    assert main(["--npd", str(planilha), "gravar"]) == 1  # não há mais nada marcado


def test_reabrir_a_mesma_cotacao_nao_incha_o_arquivo(planilha):
    """Cada reescrita da aba registrava as fotos de novo. Cinco execuções e a
    planilha carregaria cinco cópias de cada imagem."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    with zipfile.ZipFile(planilha) as z:
        depois_da_primeira = len([n for n in z.namelist() if n.startswith("xl/media/")])

    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    with zipfile.ZipFile(planilha) as z:
        depois_da_segunda = len([n for n in z.namelist() if n.startswith("xl/media/")])
        abas = re.findall(r'<sheet name="([^"]+)"', z.read("xl/workbook.xml").decode())

    assert depois_da_segunda == depois_da_primeira
    assert abas.count(NOME_ABA) == 1, "a segunda execução duplicou a aba"


# ------------------------------------------------------- recusas com motivo

def test_nome_repetido_e_recusado(planilha):
    """O nome é a chave entre as duas abas; repetido, deixa de ser chave."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    linha = PRIMEIRA_LINHA
    edicoes = {
        _celula(linha, COL_MARCAR): "x",
        _celula(linha, COL_NCM): "85166000",
        _celula(linha, COL_MARCA): "Marchesoni",
    }
    digitar_na_aba(planilha, edicoes)
    main(["--npd", str(planilha), "gravar"])

    digitar_na_aba(planilha, edicoes)  # marca o mesmo produto outra vez
    assert main(["--npd", str(planilha), "gravar"]) == 3


def test_nome_ja_existente_na_planilha_antiga_e_recusado(planilha):
    """Os 85 produtos antigos estão em `sharedStrings`; olhar só as células
    inline deixaria passar justamente o caso mais provável."""
    parametros = mod_par.ler_da_planilha(planilha)
    from npd_tool.escrita.mapeamento import LinhaPreparada

    repetido = LinhaPreparada(
        nome="Fritadeira 8000W 1 cesto 12Litros",  # já está na Priorizacao!A5
        fornecedor="Teste",
        marca="Marchesoni",
        ano=2026,
    )
    with pytest.raises(NomeRepetido) as erro:
        inserir_produtos(planilha, planilha, [repetido], parametros=parametros)
    assert "Fritadeira" in str(erro.value)


def test_marca_em_branco_para_a_gravacao_com_a_linha_apontada(planilha):
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    linha = PRIMEIRA_LINHA
    digitar_na_aba(
        planilha,
        {_celula(linha, COL_MARCAR): "x", _celula(linha, COL_NCM): "85166000"},
    )
    with pytest.raises(SelecaoInvalida) as erro:
        mod_candidatos.preparar_selecionados(
            mod_candidatos.ler_selecao(planilha),
            mod_par.ler_da_planilha(planilha),
            mod_ncm.ler_da_planilha(planilha),
        )
    assert str(linha) in str(erro.value)
    assert "Marca" in str(erro.value)


def test_cotacao_que_saiu_do_lugar_para_com_mensagem(planilha, tmp_path):
    copia = tmp_path / "cotacao_temporaria.xlsx"
    shutil.copy(FRESPRO, copia)
    main(["--npd", str(planilha), "abrir", str(copia)])
    digitar_na_aba(
        planilha,
        {
            _celula(PRIMEIRA_LINHA, COL_MARCAR): "x",
            _celula(PRIMEIRA_LINHA, COL_MARCA): "Marchesoni",
        },
    )
    copia.unlink()

    with pytest.raises(SelecaoInvalida) as erro:
        mod_candidatos.preparar_selecionados(
            mod_candidatos.ler_selecao(planilha),
            mod_par.ler_da_planilha(planilha),
            mod_ncm.ler_da_planilha(planilha),
        )
    assert "não está mais onde estava" in str(erro.value)


def test_sem_planilha_informada_o_comando_recusa(capsys):
    import os

    anterior = os.environ.pop("NPD_PLANILHA", None)
    try:
        assert main(["abrir", str(FRESPRO)]) == 2
        assert "informe a planilha" in capsys.readouterr().err
    finally:
        if anterior:
            os.environ["NPD_PLANILHA"] = anterior


def test_conferir_sem_aba_candidatos_explica_o_que_fazer(planilha, capsys):
    assert main(["--npd", str(planilha), "conferir"]) == 2
    assert "rode `abrir` primeiro" in capsys.readouterr().err


# ----------------------------------------------- a planilha continua inteira

def test_a_aba_nova_nao_estraga_o_que_ja_existia(planilha):
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])

    wb = openpyxl.load_workbook(planilha)
    try:
        assert wb.sheetnames[:6] == [
            "Funil",
            "Pesos",
            "Priorizacao",
            "Ranking",
            "Matriz",
            "Guia",
        ]
        assert wb.sheetnames[-1] == NOME_ABA  # entrou no fim, não no meio
        assert wb["Funil"]["E90"].value == "Carrinho de serviço"
        assert wb["Pesos"]["D24"].value == 5.2
        assert wb["Priorizacao"]["A5"].value == "Fritadeira 8000W 1 cesto 12Litros"
    finally:
        wb.close()

    with zipfile.ZipFile(NPD) as origem, zipfile.ZipFile(planilha) as escrita:
        antigas = {n for n in origem.namelist() if n.startswith("xl/media/")}
        assert antigas <= set(escrita.namelist())
        for nome in antigas:
            assert origem.read(nome) == escrita.read(nome)


def test_os_filtros_automaticos_continuam_apontando_para_as_abas_certas(planilha):
    """Os `definedNames` do arquivo apontam abas por índice. A aba nova tem que
    entrar no fim, ou o filtro do Funil passa a valer para outra aba."""
    main(["--npd", str(planilha), "abrir", str(FRESPRO)])
    with zipfile.ZipFile(planilha) as z:
        workbook = z.read("xl/workbook.xml").decode("utf-8")

    assert 'localSheetId="0" hidden="1">Funil!' in workbook
    assert 'localSheetId="2" hidden="1">Priorizacao!' in workbook
