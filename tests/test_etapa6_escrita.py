"""Critério de aceite da Etapa 6 (PLANO.md seção 11).

    "inserir cinco produtos de três cotações diferentes; abrir no Excel; os
     cinco aparecem nas duas abas; o score calcula; nada antigo quebrou."

Os cinco produtos vêm da Astar (PDF), da Frespro (xlsx transposto) e da Yip
(PDF) — três formatos diferentes, de propósito.

"Abrir no Excel" é o único item que um teste não alcança; para ele existe
`tests/manual_etapa6.py`, com o roteiro dos sete itens da seção 7.3. O que dá
para verificar aqui é tudo o mais, e em especial os dois modos de falha que
não dão erro nenhum na hora: o nome que não casa entre as abas, e o produto
que fica fora do intervalo do RANK.EQ.
"""
from __future__ import annotations

import hashlib
import re
import shutil
import zipfile
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from npd_tool.custo import ncm as mod_ncm
from npd_tool.custo import parametros as mod_par
from npd_tool.custo.ncm import AliquotasNCM, TabelaNCM
from npd_tool.escrita import mapeamento
from npd_tool.escrita.backup import backups_existentes, caminho_do_backup, fazer_backup
from npd_tool.escrita.mapeamento import Complementos, preparar_linha
from npd_tool.escrita.ooxml import (
    CapacidadeEsgotada,
    escrever_parametros_custo,
    estender_rank_eq,
    inserir_produtos,
    proxima_linha_livre,
)
from npd_tool.ingest.detector import ler_cotacao
from npd_tool.normalizar.nomes import trim_estavel
from npd_tool.relatorio import gravar_relatorio, montar_relatorio

FIXTURES = Path(__file__).parent / "fixtures"
NPD = FIXTURES / "NPD_2026_04_08_26.xlsx"
ASTAR = FIXTURES / "Astar~Milton Quotation.pdf"
FRESPRO = FIXTURES / "Convection Oven project Quotation from Frespro--20260713.xlsx"
JIABAO = FIXTURES / "Quotation Jiabao 2020716.pdf"

PRIMEIRA_LINHA_FUNIL = 91
PRIMEIRA_LINHA_PRIO = 90
FOTOS_ORIGINAIS = 71


@pytest.fixture(scope="module")
def npd_com_parametros(tmp_path_factory):
    """A Etapa 5 roda antes: sem os parâmetros na aba `Pesos` não há markup
    para a coluna P referenciar nem alíquota para o custo."""
    destino = tmp_path_factory.mktemp("etapa6") / "NPD_base.xlsx"

    tabela = TabelaNCM()
    tabela.adicionar(
        AliquotasNCM("85166000", "Fornos elétricos", Decimal("0.20"), Decimal("0.10"))
    )
    tabela.adicionar(
        AliquotasNCM("84185000", "Refrigeração", Decimal("0.14"), Decimal("0.05"))
    )
    escrever_parametros_custo(
        NPD, destino, mod_par.ParametrosCusto.padrao(), tabela
    )
    return destino


@pytest.fixture(scope="module")
def cinco_produtos(npd_com_parametros):
    """Cinco fichas de três cotações, em três formatos diferentes."""
    parametros = mod_par.ler_da_planilha(npd_com_parametros)
    tabela = mod_ncm.ler_da_planilha(npd_com_parametros)

    fichas = (
        ler_cotacao(ASTAR)[:2] + ler_cotacao(FRESPRO)[:2] + ler_cotacao(JIABAO)[:1]
    )
    assert len(fichas) == 5

    linhas = [
        preparar_linha(
            ficha,
            Complementos(marca="Marchesoni", ncm="85166000", unidades_no_lote=500),
            parametros,
            tabela,
        )
        for ficha in fichas
    ]
    return linhas, parametros


@pytest.fixture(scope="module")
def planilha_escrita(tmp_path_factory, npd_com_parametros, cinco_produtos):
    linhas, parametros = cinco_produtos
    destino = tmp_path_factory.mktemp("etapa6-escrita") / "NPD_com_produtos.xlsx"
    shutil.copy(npd_com_parametros, destino)
    resultado = inserir_produtos(destino, destino, linhas, parametros=parametros)
    return destino, linhas, resultado


# ------------------------------------------------- os cinco nas duas abas

def test_cinco_produtos_de_tres_cotacoes_entram_nas_duas_abas(planilha_escrita):
    destino, linhas, resultado = planilha_escrita
    assert len(resultado.produtos) == 5

    wb = openpyxl.load_workbook(destino)
    try:
        funil, prio = wb["Funil"], wb["Priorizacao"]
        for indice, (linha, escrito) in enumerate(zip(linhas, resultado.produtos)):
            assert escrito.linha_funil == PRIMEIRA_LINHA_FUNIL + indice
            assert escrito.linha_priorizacao == PRIMEIRA_LINHA_PRIO + indice
            assert funil[f"E{escrito.linha_funil}"].value == linha.nome
            assert prio[f"A{escrito.linha_priorizacao}"].value == linha.nome
            assert funil[f"G{escrito.linha_funil}"].value == "Importado"
            assert prio[f"E{escrito.linha_priorizacao}"].value == "Análise viabilidade"
    finally:
        wb.close()

    fornecedores = {linha.fornecedor for linha in linhas}
    assert len(fornecedores) == 3


def test_o_nome_casa_byte_a_byte_entre_as_abas(planilha_escrita):
    """O vínculo é o texto do nome (PLANO.md 3.5.1). Um caractere diferente e o
    score some sem erro nenhum — é a falha mais silenciosa desta etapa."""
    destino, _, resultado = planilha_escrita
    wb = openpyxl.load_workbook(destino)
    try:
        funil, prio = wb["Funil"], wb["Priorizacao"]
        nomes_prio = [prio[f"A{r}"].value for r in range(5, 131)]
        for escrito in resultado.produtos:
            nome = funil[f"E{escrito.linha_funil}"].value
            # é isto que a fórmula do Funil faz antes de procurar
            procurado = trim_estavel(nome)
            assert nome == procurado, "o nome não sobrevive ao TRIM/SUBSTITUTE"
            assert nomes_prio.count(procurado) == 1, f"{nome!r} não casou uma única vez"
    finally:
        wb.close()


def test_as_formulas_de_score_e_posicao_foram_escritas(planilha_escrita):
    destino, _, resultado = planilha_escrita
    wb = openpyxl.load_workbook(destino)
    try:
        funil = wb["Funil"]
        for escrito in resultado.produtos:
            linha = escrito.linha_funil
            score = funil[f"AB{linha}"].value
            posicao = funil[f"AC{linha}"].value
            assert score.startswith("=IFERROR(INDEX(Priorizacao!$AU$5:$AU$130")
            assert f"$E{linha}" in score
            assert "_xlfn.RANK.EQ" in posicao
            assert funil[f"K{linha}"].value.startswith("=IF(N($AB")
    finally:
        wb.close()


def test_a_linha_da_priorizacao_deixa_de_ser_oculta(planilha_escrita):
    """As vagas livres da Priorizacao vêm com `hidden="1"`; um produto em linha
    oculta pontua mas some da vista de quem precisa dar as notas."""
    destino, _, resultado = planilha_escrita
    wb = openpyxl.load_workbook(destino)
    try:
        prio = wb["Priorizacao"]
        for escrito in resultado.produtos:
            assert not prio.row_dimensions[escrito.linha_priorizacao].hidden
        assert prio.row_dimensions[120].hidden  # as ainda não usadas continuam
    finally:
        wb.close()


# ------------------------------------------------------ extensão do RANK.EQ

def test_rank_eq_estendido_ate_a_ultima_linha_escrita(planilha_escrita):
    """Armadilha 3.5.2: o intervalo fixo `$AB$2:$AB$91` deixaria de fora tudo a
    partir da segunda inserção, sem erro nenhum na tela."""
    destino, _, resultado = planilha_escrita
    assert resultado.ultima_linha_rank == 95

    with zipfile.ZipFile(destino) as z:
        funil_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    intervalos = set(re.findall(r"\$AB\$2:\$AB\$(\d+)", funil_xml))
    assert intervalos == {"95"}, "sobrou intervalo antigo em alguma fórmula"


def test_rank_eq_alcanca_as_linhas_antigas_tambem(planilha_escrita):
    """As fórmulas antigas são compartilhadas: o mestre é que carrega o texto.
    Se a extensão não pegasse nelas, o produto novo entraria no ranking mas os
    antigos continuariam ranqueando entre si."""
    destino, _, _ = planilha_escrita
    wb = openpyxl.load_workbook(destino)
    try:
        assert "$AB$2:$AB$95" in wb["Funil"]["AC2"].value
    finally:
        wb.close()


def test_estender_rank_eq_nao_encolhe_o_intervalo():
    xml = 'x<f>RANK.EQ($AB5,$AB$2:$AB$120,0)</f>'
    novo, ultima = estender_rank_eq(xml, 95)
    assert ultima == 120
    assert novo == xml


# ------------------------------------------------------- nada antigo quebrou

def _celulas(caminho, aba, primeira, ultima, ultima_coluna):
    wb = openpyxl.load_workbook(caminho)
    try:
        ws = wb[aba]
        return {
            f"{get_column_letter(c)}{r}": ws.cell(row=r, column=c).value
            for r in range(primeira, ultima + 1)
            for c in range(1, ultima_coluna + 1)
        }
    finally:
        wb.close()


@pytest.mark.parametrize(
    "aba,primeira,ultima,colunas",
    [("Funil", 1, 90, 29), ("Priorizacao", 1, 89, 51), ("Pesos", 1, 42, 7)],
)
def test_nenhuma_celula_preexistente_mudou(planilha_escrita, aba, primeira, ultima, colunas):
    """Item 7 da seção 7.3, comparado célula a célula.

    A única diferença aceita é o intervalo do RANK.EQ, que muda de propósito.
    """
    destino, _, _ = planilha_escrita
    antes = _celulas(NPD, aba, primeira, ultima, colunas)
    depois = _celulas(destino, aba, primeira, ultima, colunas)

    diferentes = {ref for ref, valor in antes.items() if depois[ref] != valor}
    for ref in diferentes:
        assert ref.startswith("AC"), f"{ref} mudou: {antes[ref]!r} -> {depois[ref]!r}"
        assert antes[ref].replace("$AB$91", "$AB$95") == depois[ref]


def test_as_71_fotos_antigas_continuam_e_as_novas_entraram(planilha_escrita):
    destino, _, resultado = planilha_escrita
    with zipfile.ZipFile(NPD) as origem, zipfile.ZipFile(destino) as escrito:
        antigas = {n for n in origem.namelist() if n.startswith("xl/media/")}
        novas = {n for n in escrito.namelist() if n.startswith("xl/media/")}
        assert antigas <= novas, "alguma foto antiga sumiu"
        assert len(antigas) == FOTOS_ORIGINAIS

        com_foto = [p for p in resultado.produtos if p.tem_foto]
        assert len(novas) == FOTOS_ORIGINAIS + len(com_foto)

        for nome in sorted(antigas):
            assert origem.read(nome) == escrito.read(nome)


def test_as_quatro_partes_do_rich_value_andam_juntas(planilha_escrita):
    """A foto na célula é uma amarração de quatro contagens. Se uma delas ficar
    para trás, a imagem aparece na linha errada — e nada acusa o erro."""
    destino, _, resultado = planilha_escrita
    com_foto = sum(1 for p in resultado.produtos if p.tem_foto)
    esperado = str(FOTOS_ORIGINAIS + com_foto)

    with zipfile.ZipFile(destino) as z:
        metadata = z.read("xl/metadata.xml").decode("utf-8")
        rdrich = z.read("xl/richData/rdrichvalue.xml").decode("utf-8")
        rels = z.read("xl/richData/richValueRel.xml").decode("utf-8")

    assert re.search(r'<valueMetadata count="(\d+)"', metadata).group(1) == esperado
    assert (
        re.search(r'<futureMetadata name="XLRICHVALUE" count="(\d+)"', metadata).group(1)
        == esperado
    )
    assert re.search(r'<rvData[^>]*count="(\d+)"', rdrich).group(1) == esperado
    assert len(re.findall(r"<rel ", rels)) == int(esperado)


def test_vm_da_celula_e_1_based(planilha_escrita):
    """O `vm` é 1-based sobre uma lista cujo índice interno é 0-based. Errar a
    base é o bug mais provável desta parte (PLANO.md 7.2) e ele não estoura."""
    destino, _, resultado = planilha_escrita
    with zipfile.ZipFile(destino) as z:
        funil_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

    primeiro = next(p for p in resultado.produtos if p.tem_foto)
    m = re.search(rf'<c r="B{primeiro.linha_funil}"[^>]*vm="(\d+)"', funil_xml)
    assert m, "a célula da foto não saiu com o atributo vm"
    assert int(m.group(1)) == FOTOS_ORIGINAIS + 1


def test_foto_jpeg_declara_o_content_type(planilha_escrita):
    """A planilha só declarava `png`; foto vinda de PDF costuma ser jpeg, e sem
    o Default o Excel abre acusando arquivo corrompido."""
    destino, _, _ = planilha_escrita
    with zipfile.ZipFile(destino) as z:
        nomes = z.namelist()
        content_types = z.read("[Content_Types].xml").decode("utf-8")

    extensoes = {n.rsplit(".", 1)[-1] for n in nomes if n.startswith("xl/media/")}
    for extensao in extensoes:
        assert f'Extension="{extensao}"' in content_types


def test_o_arquivo_de_origem_nunca_e_tocado(tmp_path, cinco_produtos):
    """Regra inegociável 1 do CLAUDE.md: ler da NPD e gravar noutro lugar não
    pode deixar marca no arquivo de origem."""
    linhas, _ = cinco_produtos
    antes = hashlib.sha256(NPD.read_bytes()).hexdigest()
    inserir_produtos(NPD, tmp_path / "saida.xlsx", linhas[:1], com_backup=False)
    assert hashlib.sha256(NPD.read_bytes()).hexdigest() == antes


# ------------------------------------------------------------------ backup

def test_gravar_por_cima_faz_backup_antes(tmp_path, npd_com_parametros, cinco_produtos):
    linhas, parametros = cinco_produtos
    alvo = tmp_path / "NPD.xlsx"
    shutil.copy(npd_com_parametros, alvo)
    antes = alvo.read_bytes()

    resultado = inserir_produtos(alvo, alvo, linhas[:2], parametros=parametros)

    assert resultado.backup is not None
    assert resultado.backup.exists()
    assert resultado.backup.read_bytes() == antes, "o backup não é o arquivo original"
    assert alvo.read_bytes() != antes
    assert backups_existentes(alvo) == [resultado.backup]


def test_backup_no_mesmo_segundo_nao_sobrescreve_o_anterior(tmp_path):
    """`abrir` seguido de `conferir` cai no mesmo segundo com facilidade. O
    segundo backup ganha sufixo — perder o primeiro não é opção, e abortar a
    gravação por causa do nome de um arquivo de segurança, tampouco."""
    origem = tmp_path / "NPD.xlsx"
    origem.write_bytes(b"primeiro conteudo")
    from datetime import datetime

    momento = datetime(2026, 8, 7, 10, 30, 0)
    primeiro = fazer_backup(origem, momento=momento)
    assert primeiro == caminho_do_backup(origem, momento=momento)

    origem.write_bytes(b"segundo conteudo")
    segundo = fazer_backup(origem, momento=momento)

    assert segundo != primeiro
    assert primeiro.read_bytes() == b"primeiro conteudo"
    assert segundo.read_bytes() == b"segundo conteudo"


# -------------------------------------------------------------- capacidade

def test_recusa_estourar_a_linha_130_da_priorizacao(tmp_path, npd_com_parametros, cinco_produtos):
    """Inserir sem estender as fórmulas gravaria produto que nunca pontua —
    pior que não inserir, porque parece que funcionou (PLANO.md 3.5.3)."""
    linhas, parametros = cinco_produtos
    alvo = tmp_path / "NPD.xlsx"
    shutil.copy(npd_com_parametros, alvo)

    demais = linhas * 10  # 50 produtos para 41 vagas
    with pytest.raises(CapacidadeEsgotada) as erro:
        inserir_produtos(alvo, alvo, demais, parametros=parametros)
    assert "130" in str(erro.value)


def test_conta_as_vagas_restantes(planilha_escrita):
    _, _, resultado = planilha_escrita
    # 130 − 94 (última escrita) = 36
    assert resultado.vagas_restantes_priorizacao == 36


def test_vaos_antigos_do_funil_sao_reportados_e_nao_usados(planilha_escrita):
    """As linhas 46, 47 e 86–88 têm fórmula mas não têm produto. A ferramenta
    não as ocupa — mexer no meio de uma lista alheia não é decisão dela — mas
    também não as esconde."""
    _, _, resultado = planilha_escrita
    assert resultado.vaos_no_funil == [46, 47, 86, 87, 88]
    assert any("46" in aviso for aviso in resultado.avisos)


def test_proxima_linha_livre_e_depois_da_ultima_ocupada():
    with zipfile.ZipFile(NPD) as z:
        funil = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        prio = z.read("xl/worksheets/sheet3.xml").decode("utf-8")
    assert proxima_linha_livre(funil, "E", 2, 992) == 91
    assert proxima_linha_livre(prio, "A", 5, 130) == 90


# --------------------------------------------- o que a ferramenta não escreve

def test_colunas_de_julgamento_humano_ficam_vazias(planilha_escrita):
    """Cód Mega vem do ERP; peças/mês, lançamento e concorrência são decisão de
    quem conhece o mercado. Preencher qualquer uma seria inventar."""
    destino, _, resultado = planilha_escrita
    wb = openpyxl.load_workbook(destino)
    try:
        funil, prio = wb["Funil"], wb["Priorizacao"]
        for escrito in resultado.produtos:
            linha = escrito.linha_funil
            for coluna in ("C", "H", "I", "L", "Q", "R", "Y", "Z", "AA"):
                assert funil[f"{coluna}{linha}"].value is None, coluna
            prio_linha = escrito.linha_priorizacao
            # notas 0–5 e peças/mês continuam com o humano
            for coluna in ("N", "R", "V", "Z", "AS"):
                assert prio[f"{coluna}{prio_linha}"].value is None, coluna
    finally:
        wb.close()


def test_g1_so_e_escrito_quando_o_humano_aceita(tmp_path, npd_com_parametros, cinco_produtos):
    """A sugestão de tensão/frequência aparece na tela; o portão zera o score,
    então gravá-la sozinha mataria produto viável em silêncio (PLANO.md 9.3)."""
    linhas, parametros = cinco_produtos
    alvo = tmp_path / "NPD.xlsx"
    shutil.copy(npd_com_parametros, alvo)

    sugerida = next((l for l in linhas if l.sugestao_g1), None)
    assert sugerida is not None, "nenhuma ficha trouxe frequência para sugerir"

    resultado = inserir_produtos(alvo, alvo, [sugerida], parametros=parametros)
    wb = openpyxl.load_workbook(alvo)
    try:
        linha = resultado.produtos[0].linha_priorizacao
        assert wb["Priorizacao"][f"F{linha}"].value is None
    finally:
        wb.close()

    aceita = mapeamento.celulas_priorizacao(
        mapeamento.LinhaPreparada(
            nome="x", fornecedor="y", marca="Marchesoni", ano=2026, g1="Passa"
        )
    )
    assert any(c.coluna == "F" and c.valor == "Passa" for c in aceita)


def test_sem_ncm_a_coluna_de_custo_fica_vazia(tmp_path, npd_com_parametros):
    """Sem NCM não há alíquota, e sem alíquota o custo não é calculado — a linha
    entra com FOB e m³ e a coluna O vazia (PLANO.md 6.6)."""
    parametros = mod_par.ler_da_planilha(npd_com_parametros)
    ficha = ler_cotacao(ASTAR)[1]
    linha = preparar_linha(
        ficha, Complementos(marca="MarcPro", ncm=None), parametros, TabelaNCM()
    )
    assert linha.custo_economico is None
    assert any("NCM" in p for p in linha.pendencias)

    alvo = tmp_path / "NPD.xlsx"
    shutil.copy(npd_com_parametros, alvo)
    resultado = inserir_produtos(alvo, alvo, [linha], parametros=parametros)

    wb = openpyxl.load_workbook(alvo)
    try:
        funil = wb["Funil"]
        numero = resultado.produtos[0].linha_funil
        assert funil[f"O{numero}"].value is None
        assert funil[f"P{numero}"].value is None  # sem custo não há revenda mínima
        assert funil[f"M{numero}"].value is not None  # FOB entra do mesmo jeito
    finally:
        wb.close()


def test_a_coluna_p_referencia_o_markup_da_aba_pesos(planilha_escrita):
    """Substitui o multiplicador digitado dentro da fórmula (PLANO.md 3.5.7)."""
    destino, linhas, resultado = planilha_escrita
    wb = openpyxl.load_workbook(destino)
    try:
        funil = wb["Funil"]
        com_custo = [
            escrito
            for linha, escrito in zip(linhas, resultado.produtos)
            if linha.custo_economico is not None
        ]
        assert com_custo, "nenhum produto teve custo calculado"
        for escrito in com_custo:
            formula = funil[f"P{escrito.linha_funil}"].value
            assert formula == f"=$O{escrito.linha_funil}*Pesos!$D$62"
            assert "2.2" not in formula and "2,2" not in formula
    finally:
        wb.close()


def test_marca_invalida_e_recusada():
    with pytest.raises(ValueError):
        Complementos(marca="Outra")


# ------------------------------------------------------------- relatório

def test_relatorio_lista_linhas_pendencias_e_capacidade(planilha_escrita, tmp_path):
    destino, linhas, resultado = planilha_escrita
    texto = montar_relatorio(linhas, resultado, destino)

    for escrito in resultado.produtos:
        assert str(escrito.linha_funil) in texto
        assert escrito.nome in texto

    assert "Pendências" in texto
    assert "Memória de cálculo" in texto
    assert "Vagas restantes" in texto
    assert str(resultado.vagas_restantes_priorizacao) in texto

    sem_m3 = [l for l in linhas if l.m3_unitario is None]
    if sem_m3:
        assert "packing list" in texto

    caminho = gravar_relatorio(texto, tmp_path)
    assert caminho.exists()
    assert caminho.read_text(encoding="utf-8") == texto


def test_relatorio_avisa_o_custo_subestimado(planilha_escrita):
    """As despesas de desembaraço por DI estão em zero enquanto o despachante
    não responde; o relatório precisa dizer isso onde o número aparece."""
    _, linhas, resultado = planilha_escrita
    texto = montar_relatorio(linhas, resultado)
    if any(l.custo_economico is not None for l in linhas):
        assert "CUSTO SUBESTIMADO" in texto
